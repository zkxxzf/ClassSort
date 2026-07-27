# -*- coding: utf-8 -*-
# ============================================================
# 版权所有 (C) 2026 @zkxxzf
# 教务分班软件 ClassSort
# 版本: 1.13 | 日期: 2026-07-28 | 作者: @zkxxzf
# ============================================================
"""Excel 导入导出 - 字段智能化识别"""
from __future__ import annotations
import os
import re
from typing import Any, Optional
from pathlib import Path

from core.models import (Student, FieldDef, RESERVED_NAME, RESERVED_PRESET,
                         FIELD_MEASURE, FIELD_CATEGORY, FIELD_TEXT)

# 字段名别名表（用于智能识别无前缀的列）
_NAME_ALIASES = {"姓名", "学生姓名", "名字", "name", "Name", "NAME"}
_PRESET_ALIASES = {"预设班级", "目标班级", "指定班级", "预设班", "preset"}
_SCORE_ALIASES = {
    "总分": ["总分", "总成绩", "总分分", "total"],
    "语文": ["语文", "语", "chinese"],
    "数学": ["数学", "数", "math"],
    "英语": ["英语", "外语", "english"],
    "物理": ["物理", "physics"],
    "化学": ["化学", "chemistry"],
    "生物": ["生物", "biology"],
    "政治": ["政治", "politics"],
    "历史": ["历史", "history"],
    "地理": ["地理", "geography"],
}
_CATEGORY_ALIASES = {
    "性别": ["性别", "sex", "gender"],
    "民族": ["民族", "nation"],
    "学校": ["学校", "毕业学校", "原学校", "来源学校", "school"],
    "类别": ["类别", "学生类别"],
}


def _read_xls_any(path: str):
    """读取 xls/xlsx 文件，返回 {sheet_name: list_of_rows}"""
    try:
        import xlrd
        wb = xlrd.open_workbook(path)
        result = {}
        for sh in wb.sheets():
            rows = []
            for r in range(sh.nrows):
                row = [sh.cell_value(r, c) for c in range(sh.ncols)]
                rows.append(row)
            result[sh.name] = rows
        return result
    except Exception:
        pass
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, data_only=True, read_only=True)
        result = {}
        for sh in wb.worksheets:
            rows = []
            for row in sh.iter_rows(values_only=True):
                rows.append(list(row))
            result[sh.title] = rows
        return result
    except Exception as e:
        raise RuntimeError(f"无法读取 Excel 文件: {e}")


def _smart_parse_column(col_name: str) -> FieldDef:
    """智能识别列字段。优先按 K/M/T 前缀，无前缀时按别名表推断。"""
    name = str(col_name).strip() if col_name is not None else ""
    # 1. 保留字段
    if name in _NAME_ALIASES or name == RESERVED_NAME:
        return FieldDef(raw_name=name, name=RESERVED_NAME, ftype="R")
    if name in _PRESET_ALIASES or name == RESERVED_PRESET:
        return FieldDef(raw_name=name, name=RESERVED_PRESET, ftype="R")
    # 2. 前缀 K/M/T
    if name and name[0].upper() in ("K", "M", "T") and len(name) > 1:
        prefix = name[0].upper()
        clean = name[1:].strip()
        return FieldDef(raw_name=name, name=clean,
                        ftype=prefix,
                        is_score=(prefix == FIELD_MEASURE),
                        is_category=(prefix == FIELD_CATEGORY))
    # 3. 别名推断
    lower = name.lower()
    for canonical, aliases in _SCORE_ALIASES.items():
        for a in aliases:
            if name == a or lower == a.lower():
                return FieldDef(raw_name=name, name=canonical,
                                ftype=FIELD_MEASURE,
                                is_score=True)
    for canonical, aliases in _CATEGORY_ALIASES.items():
        for a in aliases:
            if name == a or lower == a.lower():
                return FieldDef(raw_name=name, name=canonical,
                                ftype=FIELD_CATEGORY,
                                is_category=True)
    # 4. 默认文本
    return FieldDef(raw_name=name, name=name, ftype=FIELD_TEXT)


def _to_float(v: Any) -> Optional[float]:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def _to_str(v: Any) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    # xlrd 可能返回 float
    if s.endswith(".0"):
        try:
            f = float(s)
            if f.is_integer():
                return str(int(f))
        except ValueError:
            pass
    return s


def load_students(path: str) -> tuple[list[Student], list[FieldDef], list[dict]]:
    """从 Excel 加载学生数据

    返回: (students, fields, raw_rows)
    字段定义顺序与原始列顺序一致。
    """
    sheets = _read_xls_any(path)
    if not sheets:
        raise RuntimeError("Excel 文件无工作表")
    # 取第一个学生表（含"姓名"列的表）
    student_sheet = None
    for sh_name, rows in sheets.items():
        if not rows:
            continue
        header = [str(c).strip() if c is not None else "" for c in rows[0]]
        for h in header:
            fd = _smart_parse_column(h)
            if fd.name == RESERVED_NAME:
                student_sheet = (sh_name, rows)
                break
        if student_sheet:
            break
    if not student_sheet:
        # 没找到含姓名列的表，取第一个
        first_key = next(iter(sheets))
        student_sheet = (first_key, sheets[first_key])

    sh_name, rows = student_sheet
    if len(rows) < 2:
        return [], [], []

    header = rows[0]
    fields = [_smart_parse_column(c) for c in header]
    students: list[Student] = []
    raw_rows: list[dict] = []
    for i, row in enumerate(rows[1:], start=0):
        # 跳过全空行
        if all(v is None or v == "" for v in row):
            continue
        st = Student(sid=len(students))
        raw: dict[str, Any] = {}
        for j, fd in enumerate(fields):
            val = row[j] if j < len(row) else None
            raw[fd.raw_name] = val
            if fd.name == RESERVED_NAME:
                st.name = _to_str(val)
            elif fd.name == RESERVED_PRESET:
                pv = _to_float(val)
                if pv is not None and pv > 0:
                    st.preset_class = int(pv)
            elif fd.is_score:
                fv = _to_float(val)
                if fv is not None:
                    st.scores[fd.name] = fv
                elif fd.name in ("总分",):  # 总分允许后算
                    pass
            elif fd.is_category:
                cv = _to_str(val)
                if cv:
                    st.categories[fd.name] = cv
            else:
                tv = _to_str(val)
                if tv:
                    st.texts[fd.name] = tv
        # 若总分未给，则由单科求和
        if "总分" not in st.scores and st.scores:
            st.scores["总分"] = sum(st.scores.values())
        st.raw = raw
        students.append(st)
        raw_rows.append(raw)
    return students, fields, raw_rows


def load_teachers(path: str) -> list[dict]:
    """加载班主任表（若有）。返回 [{name, preset_class}, ...]"""
    sheets = _read_xls_any(path)
    for sh_name, rows in sheets.items():
        if not rows:
            continue
        header = [str(c).strip() if c is not None else "" for c in rows[0]]
        # 寻找含"名称"/"姓名"列的表，且行数较少（班主任表）
        name_idx = -1
        preset_idx = -1
        for j, h in enumerate(header):
            if h in ("名称", "姓名", "班主任", "老师"):
                name_idx = j
            if h in ("预设班级", "目标班级", "指定班级"):
                preset_idx = j
        if name_idx >= 0 and len(rows) <= 50:
            teachers = []
            for row in rows[1:]:
                if all(v is None or v == "" for v in row):
                    continue
                t = {"name": _to_str(row[name_idx]) if name_idx < len(row) else ""}
                if preset_idx >= 0 and preset_idx < len(row):
                    pv = _to_float(row[preset_idx])
                    t["preset_class"] = int(pv) if pv else None
                else:
                    t["preset_class"] = None
                teachers.append(t)
            return teachers
    return []


# =================== 导出 ===================
def export_classes(path: str, students: list[Student],
                   fields: list[FieldDef], classes: list,
                   teachers: Optional[list[dict]] = None):
    """导出分班结果到 Excel

    生成 3 类工作表：
    - 全体总表：所有学生 + 分班结果列
    - 各班分表：每个班级一张表
    - 班主任表（若提供）
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    # 默认表改名
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(left=Side(style="thin", color="CCCCCC"),
                    right=Side(style="thin", color="CCCCCC"),
                    top=Side(style="thin", color="CCCCCC"),
                    bottom=Side(style="thin", color="CCCCCC"))

    def write_header(ws, headers: list[str]):
        for j, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=j, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border

    def write_row(ws, row_idx: int, values: list[Any]):
        for j, v in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=j, value=v)
            cell.alignment = center
            cell.border = border

    def auto_width(ws, headers: list[str], rows: list[list[Any]]):
        for j, h in enumerate(headers, start=1):
            max_len = len(str(h))
            for r in rows:
                v = r[j - 1] if j - 1 < len(r) else ""
                max_len = max(max_len, len(str(v)))
            ws.column_dimensions[get_column_letter(j)].width = min(max_len + 4, 30)

    # 1. 全体总表
    ws = wb.active
    ws.title = "全体总表"
    headers = [fd.raw_name for fd in fields] + ["分班结果"]
    write_header(ws, headers)
    all_rows = []
    for st in students:
        row = []
        for fd in fields:
            v = st.raw.get(fd.raw_name)
            if v is None:
                v = ""
            row.append(v)
        row.append(f"{st.assigned_class + 1}班" if st.assigned_class is not None else "")
        all_rows.append(row)
    for i, r in enumerate(all_rows, start=2):
        write_row(ws, i, r)
    auto_width(ws, headers, all_rows)
    ws.freeze_panes = "A2"

    # 2. 各班分表
    for cls in classes:
        ws = wb.create_sheet(title=cls.name)
        write_header(ws, headers)
        cls_rows = []
        for st in cls.students:
            row = []
            for fd in fields:
                v = st.raw.get(fd.raw_name)
                if v is None:
                    v = ""
                row.append(v)
            row.append(cls.name)
            cls_rows.append(row)
        for i, r in enumerate(cls_rows, start=2):
            write_row(ws, i, r)
        auto_width(ws, headers, cls_rows)
        ws.freeze_panes = "A2"

    # 3. 班级统计表
    ws = wb.create_sheet(title="班级统计")
    stat_headers = ["班级", "人数"]
    # 加上各分数字段均分
    score_fields = [fd.name for fd in fields if fd.is_score]
    stat_headers += [f"{f}均分" for f in score_fields]
    # 加上各类别分布
    cat_fields = [fd.name for fd in fields if fd.is_category]
    cat_value_set: dict[str, set[str]] = {f: set() for f in cat_fields}
    for st in students:
        for f in cat_fields:
            v = st.category_of(f)
            if v:
                cat_value_set[f].add(v)
    for f in cat_fields:
        for v in sorted(cat_value_set[f]):
            stat_headers.append(f"{f}_{v}")
    write_header(ws, stat_headers)
    for i, cls in enumerate(classes, start=2):
        row = [cls.name, len(cls.students)]
        # 均分
        for f in score_fields:
            if cls.students:
                avg = sum(s.score_of(f) for s in cls.students) / len(cls.students)
                row.append(round(avg, 2))
            else:
                row.append(0)
        # 类别分布
        for f in cat_fields:
            cnt: dict[str, int] = {v: 0 for v in cat_value_set[f]}
            for s in cls.students:
                v = s.category_of(f)
                if v in cnt:
                    cnt[v] += 1
            for v in sorted(cat_value_set[f]):
                row.append(cnt[v])
        write_row(ws, i, row)
    auto_width(ws, stat_headers, [])
    ws.freeze_panes = "A2"

    # 4. 班主任表（若有）
    if teachers:
        ws = wb.create_sheet(title="班主任")
        th = ["名称", "预设班级"]
        write_header(ws, th)
        for i, t in enumerate(teachers, start=2):
            pc = t.get("preset_class")
            write_row(ws, i, [t.get("name", ""), pc if pc else ""])
        auto_width(ws, th, [])

    # 保存
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xls":
        path = os.path.splitext(path)[0] + ".xlsx"
    wb.save(path)
    return path
