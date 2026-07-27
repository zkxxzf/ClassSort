# -*- coding: utf-8 -*-
# ============================================================
# 版权所有 (C) 2026 @zkxxzf
# 教务分班软件 ClassSort
# 版本: 1.13 | 日期: 2026-07-28 | 作者: @zkxxzf
# ============================================================
"""教务分班软件 - GUI 主窗口

Tab 布局:
1. 数据导入 - Excel 选择 / 字段识别 / 异常检测
2. 约束设置 - 捆绑/互斥/分段/班型/班级人数
3. 分班配置 - 权重/容差/开关
4. 执行分班 - 运行/进度/统计
5. 手动调班 - 一对一/多对多换班
6. 结果导出 - 导出 Excel
"""
from __future__ import annotations
import os
import sys
import threading
import tkinter as tk
from datetime import datetime
from tkinter import ttk, filedialog, messagebox, simpledialog
from typing import Optional

# 允许从 app 目录运行
_HERE = os.path.dirname(os.path.abspath(__file__))
if _HERE not in sys.path:
    sys.path.insert(0, _HERE)

from core.models import Student, TaskConfig, FieldDef, ClassTypeSpec, SegmentRule
from core.constraints import ConstraintManager
from core.engine import BalanceEngine, EngineStats
from adjust import ManualAdjuster, detect_anomalies
from dataio.excel_handler import load_students, load_teachers, export_classes

# 版本信息
APP_VERSION = "1.13"
APP_AUTHOR = "@zkxxzf"
APP_DATE = "2026-07-27"


def _date_seed() -> int:
    """根据当天日期生成随机种子（YYYYMMDD）"""
    return int(datetime.now().strftime("%Y%m%d"))


def _resolve_icon() -> Optional[str]:
    """解析图标文件路径，兼容开发模式与 PyInstaller 打包模式"""
    base = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
    for name in ('ioc.ico', 'ioc.png'):
        p = os.path.join(base, name)
        if os.path.isfile(p):
            return p
    dev_p = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'ioc.ico')
    if os.path.isfile(dev_p):
        return dev_p
    return None


class App:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title(f"教务分班软件 v{APP_VERSION}")
        self.root.geometry("1180x760")
        self.root.minsize(1000, 660)
        icon_path = _resolve_icon()
        if icon_path and icon_path.endswith('.ico'):
            try:
                self.root.iconbitmap(icon_path)
            except Exception:
                pass
        elif icon_path:
            try:
                img = tk.PhotoImage(file=icon_path)
                self.root.iconphoto(True, img)
                self.root._icon_ref = img
            except Exception:
                pass

        # 状态
        self.students: list[Student] = []
        self.fields: list[FieldDef] = []
        self.teachers: list[dict] = []
        self.config = TaskConfig()
        self.constraints = ConstraintManager()
        self.engine: Optional[BalanceEngine] = None
        self.adjuster: Optional[ManualAdjuster] = None
        self.classes: list = []
        self.stats: Optional[EngineStats] = None
        self.excel_path: str = ""
        self._cancel_flag = False
        # 手动调班日志相关
        self.auto_stats_text: Optional[str] = None  # 自动分班统计快照
        self.swap_log: list[str] = []                # 手动换班日志条目

        self._setup_style()
        self._build_ui()
        self._build_menu()

    # =================== UI 构建 ===================
    def _setup_style(self):
        """配置全局 ttk 样式"""
        style = ttk.Style()
        try:
            style.theme_use("clam")
        except Exception:
            pass
        # 字体
        default_font = ("Microsoft YaHei UI", 10)
        style.configure(".", font=default_font)
        style.configure("TLabel", font=default_font)
        style.configure("TButton", font=default_font, padding=4)
        style.configure("TLabelframe.Label", font=("Microsoft YaHei UI", 10, "bold"))
        style.configure("Treeview", font=("Microsoft YaHei UI", 9), rowheight=24)
        style.configure("Treeview.Heading", font=("Microsoft YaHei UI", 9, "bold"))
        # 颜色
        style.configure("Accent.TButton", foreground="#0066cc")
        style.map("TButton",
                  foreground=[("active", "#0066cc")],
                  background=[("active", "#e6f0ff")])

    def _build_menu(self):
        """构建顶部菜单栏"""
        menubar = tk.Menu(self.root)
        # 文件菜单
        file_menu = tk.Menu(menubar, tearoff=False)
        file_menu.add_command(label="导入 Excel…", command=self._on_pick_excel, accelerator="Ctrl+O")
        file_menu.add_command(label="导出 Excel…", command=self._on_export, accelerator="Ctrl+E")
        file_menu.add_separator()
        file_menu.add_command(label="退出", command=self.root.quit)
        menubar.add_cascade(label="文件", menu=file_menu)
        # 帮助菜单
        help_menu = tk.Menu(menubar, tearoff=False)
        help_menu.add_command(label="关于", command=self._show_about)
        menubar.add_cascade(label="帮助", menu=help_menu)
        self.root.config(menu=menubar)
        # 快捷键绑定
        self.root.bind("<Control-o>", lambda e: self._on_pick_excel())
        self.root.bind("<Control-O>", lambda e: self._on_pick_excel())
        self.root.bind("<Control-e>", lambda e: self._on_export())
        self.root.bind("<Control-E>", lambda e: self._on_export())

    def _make_scrollable(self, parent):
        """将 parent 内容包装为可滚动区域，返回内部 Frame

        双重绑定：
        - Canvas <Configure>: 同步内层 Frame 宽度（防止内容被裁剪）
        - Inner  <Configure>: 更新 scrollregion（内容变化时滚动区域同步）
        """
        canvas = tk.Canvas(parent, highlightthickness=0, bg="#f0f0f0")
        vsb = ttk.Scrollbar(parent, orient="vertical", command=canvas.yview)
        inner = ttk.Frame(canvas)
        inner_id = canvas.create_window((0, 0), window=inner, anchor="nw")

        def _on_canvas_configure(_e):
            w = canvas.winfo_width()
            if w > 0:
                canvas.itemconfig(inner_id, width=w)
            canvas.configure(scrollregion=canvas.bbox("all"))
        canvas.bind("<Configure>", _on_canvas_configure)

        def _on_inner_configure(_e):
            canvas.configure(scrollregion=canvas.bbox("all"))
        inner.bind("<Configure>", _on_inner_configure)

        canvas.configure(yscrollcommand=vsb.set)

        canvas.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        def _on_wheel(e):
            canvas.yview_scroll(int(-1 * (e.delta / 120)), "units")
        canvas.bind("<Enter>", lambda e: canvas.bind_all("<MouseWheel>", _on_wheel))
        canvas.bind("<Leave>", lambda e: canvas.unbind_all("<MouseWheel>"))
        # 保存 canvas 引用，便于手动刷新 scrollregion
        inner._scroll_canvas = canvas
        return inner

    def _update_scrollregion(self, frame):
        """显式刷新可滚动区域的 scrollregion（数据填充后调用）"""
        if hasattr(frame, "_scroll_canvas"):
            self.root.update_idletasks()
            frame._scroll_canvas.configure(scrollregion=frame._scroll_canvas.bbox("all"))

    def _show_about(self):
        messagebox.showinfo("关于",
            f"教务分班软件 v{APP_VERSION}\n"
            f"作者: {APP_AUTHOR}\n"
            f"日期: {APP_DATE}\n\n"
            "功能：Excel 导入导出 / 智能字段识别 / 总分单科均分均衡 /\n"
            "      分段均衡 / 类别均衡 / 捆绑互斥 / 手动调班 / 异常检测 /\n"
            "      重点班/普通班分层分班\n\n"
            "完全单机运行，无需联网。")

    def _build_ui(self):
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill="both", expand=True, padx=8, pady=8)

        self.tab_import = ttk.Frame(self.notebook)
        self.tab_constraint = ttk.Frame(self.notebook)
        self.tab_config = ttk.Frame(self.notebook)
        self.tab_run = ttk.Frame(self.notebook)
        self.tab_adjust = ttk.Frame(self.notebook)
        self.tab_stats = ttk.Frame(self.notebook)
        self.tab_export = ttk.Frame(self.notebook)

        self.notebook.add(self.tab_import, text="1. 数据导入")
        self.notebook.add(self.tab_constraint, text="2. 约束设置")
        self.notebook.add(self.tab_config, text="3. 分班配置")
        self.notebook.add(self.tab_run, text="4. 执行分班")
        self.notebook.add(self.tab_adjust, text="5. 手动调班")
        self.notebook.add(self.tab_stats, text="6. 班级统计")
        self.notebook.add(self.tab_export, text="7. 结果导出")

        self._build_import_tab()
        self._build_constraint_tab()
        self._build_config_tab()
        self._build_run_tab()
        self._build_adjust_tab()
        self._build_stats_tab()
        self._build_export_tab()

        # 切换标签页时自动刷新数据（手动调班/结果导出页进入即显示）
        self.notebook.bind("<<NotebookTabChanged>>", self._on_tab_changed)

        # 状态栏
        self.status_var = tk.StringVar(value="就绪")
        sb = ttk.Label(self.root, textvariable=self.status_var,
                       relief="sunken", anchor="w")
        sb.pack(side="bottom", fill="x")

    def _on_tab_changed(self, _event):
        """切换标签页时自动刷新对应页面的数据"""
        try:
            current = self.notebook.index(self.notebook.select())
        except Exception:
            return
        # 5. 手动调班页：进入时自动加载分班结果
        if current == 4 and self.classes:
            if not self.class_a_var.get() or not self.list_a.get_children():
                self._refresh_adjust_tab()
        # 6. 班级统计页：进入时自动刷新统计
        elif current == 5 and self.classes:
            self._refresh_stats_tab()
        # 7. 结果导出页：进入时自动显示预览
        elif current == 6 and self.students and self.classes:
            self._refresh_export_preview()

    # ---------- 1. 数据导入 ----------
    def _build_import_tab(self):
        f = self.tab_import
        top = ttk.LabelFrame(f, text="Excel 名单导入")
        top.pack(fill="x", padx=8, pady=6)

        row = ttk.Frame(top); row.pack(fill="x", padx=6, pady=6)
        ttk.Button(row, text="选择 Excel 文件…",
                   command=self._on_pick_excel).pack(side="left")
        self.path_var = tk.StringVar(value="未选择")
        ttk.Label(row, textvariable=self.path_var).pack(
            side="left", padx=10, fill="x", expand=True)

        # 配置行
        row2 = ttk.Frame(top); row2.pack(fill="x", padx=6, pady=4)
        ttk.Label(row2, text="班级数:").pack(side="left")
        self.num_classes_var = tk.IntVar(value=5)
        ttk.Spinbox(row2, from_=2, to=50, width=5,
                    textvariable=self.num_classes_var).pack(side="left", padx=4)
        ttk.Label(row2, text="随机种子:").pack(side="left", padx=(16, 0))
        self.seed_var = tk.IntVar(value=_date_seed())
        ttk.Spinbox(row2, from_=1, to=99999999, width=10,
                    textvariable=self.seed_var).pack(side="left", padx=4)
        ttk.Button(row2, text="📅 今日种子",
                   command=lambda: self.seed_var.set(_date_seed())).pack(side="left", padx=4)
        ttk.Button(row2, text="检测异常",
                   command=self._on_detect_anomaly).pack(side="right")

        # 字段预览
        mid = ttk.LabelFrame(f, text="字段识别结果")
        mid.pack(fill="both", expand=True, padx=8, pady=4)
        cols = ("序号", "原始列名", "字段名", "类型", "说明")
        self.field_tree = ttk.Treeview(mid, columns=cols, show="headings", height=6)
        for c in cols:
            self.field_tree.heading(c, text=c)
            self.field_tree.column(c, width=120, anchor="center")
        self.field_tree.pack(fill="both", expand=True, padx=4, pady=4)

        # 学生预览
        bot = ttk.LabelFrame(f, text="学生数据预览（前 20 条）")
        bot.pack(fill="both", expand=True, padx=8, pady=4)
        self.stu_tree = ttk.Treeview(bot, show="headings", height=8)
        self.stu_tree.pack(fill="both", expand=True, padx=4, pady=4)

        # 异常报告
        self.anomaly_text = tk.Text(f, height=4, wrap="word")
        self.anomaly_text.pack(fill="x", padx=8, pady=4)

    def _on_pick_excel(self):
        path = filedialog.askopenfilename(
            title="选择分班名单 Excel",
            filetypes=[("Excel 文件", "*.xls *.xlsx"), ("所有文件", "*.*")])
        if not path:
            return
        try:
            students, fields, _ = load_students(path)
            teachers = load_teachers(path)
        except Exception as e:
            messagebox.showerror("导入失败", str(e))
            return
        self.students = students
        self.fields = fields
        self.teachers = teachers
        self.excel_path = path
        self.path_var.set(f"{os.path.basename(path)}  （{len(students)} 名学生，"
                          f"{len(fields)} 个字段，{len(teachers)} 名班主任）")
        self._refresh_field_tree()
        self._refresh_stu_tree()
        self.status_var.set(f"已加载 {len(students)} 名学生")

    def _refresh_field_tree(self):
        self.field_tree.delete(*self.field_tree.get_children())
        type_desc = {"K": "类别(人数均衡)", "M": "分数(均分均衡)",
                     "T": "文本", "R": "保留字段"}
        for i, fd in enumerate(self.fields, start=1):
            self.field_tree.insert("", "end", values=(
                i, fd.raw_name, fd.name, fd.ftype, type_desc.get(fd.ftype, "")))

    def _refresh_stu_tree(self):
        for c in self.stu_tree.get_children():
            self.stu_tree.delete(c)
        self.stu_tree["columns"] = [fd.raw_name for fd in self.fields] or ["empty"]
        for fd in self.fields:
            self.stu_tree.heading(fd.raw_name, text=fd.raw_name)
            self.stu_tree.column(fd.raw_name, width=80, anchor="center")
        if not self.fields:
            self.stu_tree.heading("empty", text="(无数据)")
        for st in self.students[:20]:
            vals = [st.raw.get(fd.raw_name, "") for fd in self.fields]
            self.stu_tree.insert("", "end", values=vals)

    def _on_detect_anomaly(self):
        if not self.students:
            messagebox.showwarning("提示", "请先导入数据")
            return
        score_fields = [fd.name for fd in self.fields if fd.is_score]
        rep = detect_anomalies(self.students, score_fields)
        self.anomaly_text.delete("1.0", "end")
        if not rep.anomalies:
            self.anomaly_text.insert("end", "✓ 未检测到异常\n")
        else:
            self.anomaly_text.insert("end",
                                     f"⚠ 检测到 {len(rep.anomalies)} 项异常:\n")
            for a in rep.anomalies:
                self.anomaly_text.insert("end", f"  • {a}\n")

    # ---------- 2. 约束设置 ----------
    def _build_constraint_tab(self):
        f = self._make_scrollable(self.tab_constraint)
        # 捆绑组
        bf = ttk.LabelFrame(f, text="同班捆绑组（学伴/双胞胎等）")
        bf.pack(fill="both", expand=True, padx=8, pady=4)
        bf_row = ttk.Frame(bf); bf_row.pack(fill="x", padx=4, pady=4)
        ttk.Button(bf_row, text="+ 添加捆绑组",
                   command=lambda: self._add_group("binding")).pack(side="left")
        ttk.Button(bf_row, text="删除选中",
                   command=lambda: self._del_group("binding")).pack(side="left", padx=4)
        self.bind_tree = ttk.Treeview(bf, columns=("序号/姓名", "目标班"),
                                       show="headings", height=5)
        self.bind_tree.heading("序号/姓名", text="学生（序号:姓名）")
        self.bind_tree.heading("目标班", text="目标班（可空）")
        self.bind_tree.pack(fill="both", expand=True, padx=4, pady=4)

        # 互斥组
        mf = ttk.LabelFrame(f, text="互斥分散组（早恋/不良小团体等）")
        mf.pack(fill="both", expand=True, padx=8, pady=4)
        mf_row = ttk.Frame(mf); mf_row.pack(fill="x", padx=4, pady=4)
        ttk.Button(mf_row, text="+ 添加互斥组",
                   command=lambda: self._add_group("mutex")).pack(side="left")
        ttk.Button(mf_row, text="删除选中",
                   command=lambda: self._del_group("mutex")).pack(side="left", padx=4)
        self.mutex_tree = ttk.Treeview(mf, columns=("学生",), show="headings", height=5)
        self.mutex_tree.heading("学生", text="学生（序号:姓名）")
        self.mutex_tree.pack(fill="both", expand=True, padx=4, pady=4)

        # 自定义分段
        sf = ttk.LabelFrame(f, text="自定义分数段均衡规则")
        sf.pack(fill="both", expand=True, padx=8, pady=4)
        sf_row = ttk.Frame(sf); sf_row.pack(fill="x", padx=4, pady=4)
        ttk.Label(sf_row, text="字段:").pack(side="left")
        self.seg_field_var = tk.StringVar(value="总分")
        ttk.Entry(sf_row, textvariable=self.seg_field_var, width=12).pack(side="left", padx=4)
        ttk.Label(sf_row, text="分段方式:").pack(side="left", padx=(8, 0))
        self.seg_by_rank_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(sf_row, text="按排名", variable=self.seg_by_rank_var).pack(side="left")
        ttk.Label(sf_row, text="切点(逗号分隔):").pack(side="left", padx=(8, 0))
        self.seg_cuts_var = tk.StringVar()
        ttk.Entry(sf_row, textvariable=self.seg_cuts_var, width=20).pack(side="left", padx=4)
        ttk.Button(sf_row, text="添加", command=self._add_segment).pack(side="left", padx=4)
        ttk.Button(sf_row, text="删除选中",
                   command=self._del_segment).pack(side="left")
        self.seg_tree = ttk.Treeview(sf, columns=("字段", "方式", "切点"),
                                      show="headings", height=4)
        self.seg_tree.heading("字段", text="字段")
        self.seg_tree.heading("方式", text="方式")
        self.seg_tree.heading("切点", text="切点")
        self.seg_tree.pack(fill="both", expand=True, padx=4, pady=4)

        # 班型
        cf = ttk.LabelFrame(f, text="班型设置（重点班/卓越班/普通班等）")
        cf.pack(fill="both", expand=True, padx=8, pady=4)
        cf_row = ttk.Frame(cf); cf_row.pack(fill="x", padx=4, pady=4)
        ttk.Label(cf_row, text="班型名:").pack(side="left")
        self.ct_name_var = tk.StringVar(value="重点班")
        ttk.Entry(cf_row, textvariable=self.ct_name_var, width=10).pack(side="left", padx=4)
        ttk.Label(cf_row, text="班级(1-based,逗号分隔):").pack(side="left", padx=(8, 0))
        self.ct_idx_var = tk.StringVar(value="1,2")
        ttk.Entry(cf_row, textvariable=self.ct_idx_var, width=10).pack(side="left", padx=4)
        ttk.Button(cf_row, text="添加", command=self._add_class_type).pack(side="left", padx=4)
        ttk.Button(cf_row, text="删除选中", command=self._del_class_type).pack(side="left")
        # 第二行：分数范围
        cf_row2 = ttk.Frame(cf); cf_row2.pack(fill="x", padx=4, pady=2)
        ttk.Label(cf_row2, text="最低分:").pack(side="left")
        self.ct_floor_var = tk.StringVar()
        ttk.Entry(cf_row2, textvariable=self.ct_floor_var, width=8).pack(side="left", padx=4)
        ttk.Label(cf_row2, text="最高分:").pack(side="left", padx=(8, 0))
        self.ct_ceil_var = tk.StringVar()
        ttk.Entry(cf_row2, textvariable=self.ct_ceil_var, width=8).pack(side="left", padx=4)
        # 第三行：名次范围
        cf_row3 = ttk.Frame(cf); cf_row3.pack(fill="x", padx=4, pady=2)
        ttk.Label(cf_row3, text="从第:").pack(side="left")
        self.ct_rank_floor_var = tk.StringVar()
        ttk.Entry(cf_row3, textvariable=self.ct_rank_floor_var, width=6).pack(side="left", padx=2)
        ttk.Label(cf_row3, text="名~").pack(side="left")
        ttk.Label(cf_row3, text="第:").pack(side="left", padx=(8, 0))
        self.ct_rank_ceil_var = tk.StringVar()
        ttk.Entry(cf_row3, textvariable=self.ct_rank_ceil_var, width=6).pack(side="left", padx=2)
        ttk.Label(cf_row3, text="名").pack(side="left")
        ttk.Label(cf_row3, text="(按总分年级排名)").pack(side="left", padx=8)
        self.ct_tree = ttk.Treeview(cf, columns=("班型", "班级", "最低分", "最高分", "名次范围"),
                                     show="headings", height=4)
        for h, w in (("班型", 80), ("班级", 80), ("最低分", 80), ("最高分", 80), ("名次范围", 100)):
            self.ct_tree.heading(h, text=h)
            self.ct_tree.column(h, width=w, anchor="center")
        self.ct_tree.pack(fill="both", expand=True, padx=4, pady=4)

        # 自定义班级人数
        sz = ttk.LabelFrame(f, text="自定义班级人数（留空=自动均分）")
        sz.pack(fill="x", padx=8, pady=4)
        self.size_frame = ttk.Frame(sz)
        self.size_frame.pack(fill="x", padx=4, pady=4)
        self.size_vars: list[tk.StringVar] = []
        ttk.Button(sz, text="刷新班级人数输入框",
                   command=self._refresh_size_inputs).pack(padx=4, pady=4)

    def _refresh_size_inputs(self):
        for w in self.size_frame.winfo_children():
            w.destroy()
        self.size_vars.clear()
        n = self.num_classes_var.get()
        for i in range(n):
            ttk.Label(self.size_frame, text=f"{i+1}班:").grid(row=0, column=i*2, padx=2)
            v = tk.StringVar()
            e = ttk.Entry(self.size_frame, textvariable=v, width=6)
            e.grid(row=0, column=i*2+1, padx=2)
            self.size_vars.append(v)

    def _pick_students_dialog(self, title: str) -> Optional[list[int]]:
        """弹出对话框让学生选择多个学生，返回 sid 列表"""
        if not self.students:
            messagebox.showwarning("提示", "请先导入学生数据")
            return None
        win = tk.Toplevel(self.root)
        win.title(title)
        win.geometry("560x520")
        win.transient(self.root)
        win.grab_set()
        result: list[int] = []

        top = ttk.Frame(win); top.pack(fill="x", padx=6, pady=6)
        ttk.Label(top, text="搜索:").pack(side="left")
        search_var = tk.StringVar()
        entry = ttk.Entry(top, textvariable=search_var, width=20)
        entry.pack(side="left", padx=4)
        entry.focus_set()
        ttk.Label(top, text="(双击行选中/取消)").pack(side="left", padx=10)

        cols = ("sid", "name", "scores", "preset")
        tree = ttk.Treeview(win, columns=cols, show="headings", height=15)
        tree.heading("sid", text="序号")
        tree.heading("name", text="姓名")
        tree.heading("scores", text="总分")
        tree.heading("preset", text="预设班")
        tree.column("sid", width=60, anchor="center")
        tree.column("name", width=120, anchor="center")
        tree.column("scores", width=100, anchor="center")
        tree.column("preset", width=80, anchor="center")

        def populate(filter_text=""):
            tree.delete(*tree.get_children())
            for st in self.students:
                if filter_text and filter_text not in st.name:
                    continue
                tree.insert("", "end", iid=str(st.sid), values=(
                    st.sid + 1, st.name, round(st.total_score, 1),
                    st.preset_class or ""))
        populate()

        # 先 pack 底部按钮栏（side=bottom），再 pack tree（expand=True），
        # 保证按钮始终可见，不会被 tree 挤出窗口。
        bot = ttk.Frame(win); bot.pack(side="bottom", fill="x", padx=6, pady=6)
        sel_label = ttk.Label(bot, text="已选: 0")
        sel_label.pack(side="left")
        ttk.Button(bot, text="取消", command=lambda: win.destroy()).pack(side="right", padx=4)
        ttk.Button(bot, text="✓ 确定", command=lambda: _on_ok()).pack(side="right", padx=4)

        tree.pack(fill="both", expand=True, padx=6, pady=4)

        def on_search(*_):
            populate(search_var.get().strip())
        search_var.trace_add("write", on_search)

        def toggle(evt):
            item = tree.identify_row(evt.y)
            if not item:
                return
            tags = tree.item(item, "tags")
            if "sel" in tags:
                tree.item(item, tags=[])
            else:
                tree.item(item, tags=["sel"])
                tree.tag_configure("sel", background="#cce5ff")
            sel = [int(i) for i in tree.get_children("") if "sel" in tree.item(i, "tags")]
            sel_label.config(text=f"已选: {len(sel)}")
        tree.bind("<Double-Button-1>", toggle)

        def _on_ok():
            for i in tree.get_children(""):
                if "sel" in tree.item(i, "tags"):
                    result.append(int(i))
            win.destroy()

        win.wait_window()
        return result if result else None

    def _add_group(self, kind: str):
        title = "添加捆绑组" if kind == "binding" else "添加互斥组"
        sids = self._pick_students_dialog(title + " - 选择学生")
        if not sids or len(sids) < 2:
            if sids:
                messagebox.showwarning("提示", "至少选择 2 名学生")
            return
        target = None
        if kind == "binding":
            # 自定义弹窗：允许留空（不指定目标班）
            target = self._ask_target_class_dialog()
            if target == "CANCEL":
                return  # 用户取消
            # target 为 None 表示不指定，整数表示 0-based 班级索引
        names = ";".join(f"{s+1}:{self.students[s].name}" for s in sids)
        if kind == "binding":
            self.constraints.add_binding(sids, target)
            self.bind_tree.insert("", "end", values=(
                names, (target + 1) if target is not None else ""))
        else:
            self.constraints.add_mutex(sids)
            self.mutex_tree.insert("", "end", values=(names,))

    def _ask_target_class_dialog(self):
        """弹出目标班选择对话框，允许留空（不指定）。
        返回值：
          - None: 不指定目标班
          - int: 0-based 班级索引
          - "CANCEL": 用户取消
        """
        win = tk.Toplevel(self.root)
        win.title("目标班（可选）")
        win.geometry("360x180")
        win.transient(self.root)
        win.grab_set()
        result = ["CANCEL"]  # 默认取消

        ttk.Label(win, text="是否指定目标班？").pack(pady=(12, 4))
        ttk.Label(win, text=f"(留空或选 0 = 不指定，1~{self.num_classes_var.get()} = 指定班)",
                  foreground="gray").pack(pady=(0, 8))

        var = tk.IntVar(value=0)
        spin = ttk.Spinbox(win, from_=0, to=self.num_classes_var.get(),
                           textvariable=var, width=8)
        spin.pack(pady=4)
        spin.focus_set()

        bot = ttk.Frame(win); bot.pack(side="bottom", fill="x", padx=8, pady=8)
        def on_cancel():
            win.destroy()
        def on_ok():
            v = var.get()
            if v == 0:
                result[0] = None  # 不指定
            else:
                result[0] = v - 1  # 转 0-based
            win.destroy()
        ttk.Button(bot, text="取消", command=on_cancel).pack(side="right", padx=4)
        ttk.Button(bot, text="✓ 确定", command=on_ok).pack(side="right", padx=4)

        win.wait_window()
        return result[0]

    def _del_group(self, kind: str):
        tree = self.bind_tree if kind == "binding" else self.mutex_tree
        sel = tree.selection()
        if not sel:
            return
        # 简单实现：清空并重建
        if kind == "binding":
            self.constraints.bindings.clear()
            self.constraints._binding_of.clear()
            tree.delete(*tree.get_children())
        else:
            self.constraints.mutexes.clear()
            self.constraints._mutex_of.clear()
            tree.delete(*tree.get_children())
        messagebox.showinfo("提示", "已清空全部组（请重新添加需要保留的）")

    def _add_segment(self):
        try:
            cuts = [float(x.strip()) for x in self.seg_cuts_var.get().split(",") if x.strip()]
        except ValueError:
            messagebox.showerror("错误", "切点必须是数字，逗号分隔")
            return
        rule = SegmentRule(field_name=self.seg_field_var.get(),
                           by_rank=self.seg_by_rank_var.get(), cuts=cuts)
        self.config.segment_rules.append(rule)
        self.seg_tree.insert("", "end", values=(
            rule.field_name, "排名" if rule.by_rank else "分数", str(cuts)))

    def _del_segment(self):
        sel = self.seg_tree.selection()
        if not sel:
            return
        idx = self.seg_tree.index(sel[0])
        self.seg_tree.delete(sel)
        if idx < len(self.config.segment_rules):
            self.config.segment_rules.pop(idx)

    def _add_class_type(self):
        try:
            idx_list = [int(x.strip()) - 1 for x in self.ct_idx_var.get().split(",") if x.strip()]
        except ValueError:
            messagebox.showerror("错误", "班级索引必须是数字，逗号分隔")
            return
        floor = float(self.ct_floor_var.get()) if self.ct_floor_var.get().strip() else None
        ceil = float(self.ct_ceil_var.get()) if self.ct_ceil_var.get().strip() else None
        r_floor = int(self.ct_rank_floor_var.get()) if self.ct_rank_floor_var.get().strip() else None
        r_ceil = int(self.ct_rank_ceil_var.get()) if self.ct_rank_ceil_var.get().strip() else None
        spec = ClassTypeSpec(name=self.ct_name_var.get(),
                             class_indices=idx_list,
                             score_floor=floor, score_ceil=ceil,
                             rank_floor=r_floor, rank_ceil=r_ceil)
        self.config.class_types.append(spec)
        rank_range = ""
        if r_floor is not None and r_ceil is not None:
            rank_range = f"第{r_floor}~{r_ceil}名"
        elif r_floor is not None:
            rank_range = f"第{r_floor}名及以后"
        elif r_ceil is not None:
            rank_range = f"前{r_ceil}名"
        self.ct_tree.insert("", "end", values=(
            spec.name, self.ct_idx_var.get(),
            floor if floor is not None else "",
            ceil if ceil is not None else "",
            rank_range))

    def _del_class_type(self):
        sel = self.ct_tree.selection()
        if not sel:
            return
        idx = self.ct_tree.index(sel[0])
        self.ct_tree.delete(sel)
        if idx < len(self.config.class_types):
            self.config.class_types.pop(idx)

    # ---------- 3. 分班配置 ----------
    def _build_config_tab(self):
        f = self._make_scrollable(self.tab_config)
        frame = ttk.Frame(f)
        frame.pack(fill="both", expand=True, padx=12, pady=8)

        # 权重
        wf = ttk.LabelFrame(frame, text="均衡权重（越大越重要）")
        wf.pack(fill="x", pady=4)
        self.weight_vars: dict[str, tk.DoubleVar] = {}
        weights = [
            ("w_total", "总分均分", 200.0),
            ("w_subject", "单科均分", 150.0),
            ("w_segment", "分段人数", 60.0),
            ("w_category", "类别人数", 50.0),
            ("w_pointwise", "逐点均衡", 30.0),
            ("w_size", "班级人数", 80.0),
        ]
        for i, (k, label, default) in enumerate(weights):
            r, c = divmod(i, 2)
            cell = ttk.Frame(wf); cell.grid(row=r, column=c, padx=8, pady=4, sticky="w")
            ttk.Label(cell, text=label + ":").pack(side="left")
            v = tk.DoubleVar(value=default)
            self.weight_vars[k] = v
            ttk.Spinbox(cell, from_=0, to=500, increment=5, width=8,
                        textvariable=v).pack(side="left", padx=4)

        # 容差
        tf = ttk.LabelFrame(frame, text="容差与精度")
        tf.pack(fill="x", pady=4)
        r = ttk.Frame(tf); r.pack(fill="x", padx=6, pady=4)
        ttk.Label(r, text="均分偏差目标:").pack(side="left")
        self.tol_mean_var = tk.DoubleVar(value=0.01)
        ttk.Spinbox(r, from_=0.001, to=10, increment=0.01, width=8,
                    textvariable=self.tol_mean_var).pack(side="left", padx=4)
        ttk.Label(r, text="人数偏差上限:").pack(side="left", padx=(16, 0))
        self.tol_count_var = tk.IntVar(value=1)
        ttk.Spinbox(r, from_=0, to=10, width=5,
                    textvariable=self.tol_count_var).pack(side="left", padx=4)

        # 开关
        sf = ttk.LabelFrame(frame, text="功能开关")
        sf.pack(fill="x", pady=4)
        self.pointwise_var = tk.BooleanVar(value=True)
        self.pinyin_var = tk.BooleanVar(value=True)
        self.seg_auto_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(sf, text="超强逐点均衡（高低分逐点公平）",
                        variable=self.pointwise_var).pack(anchor="w", padx=6)
        ttk.Checkbutton(sf, text="姓名同音/相近分散",
                        variable=self.pinyin_var).pack(anchor="w", padx=6)
        ttk.Checkbutton(sf, text="自动分段均衡（按名次前1/3、中1/3、后1/3三段均衡，无需设置）",
                        variable=self.seg_auto_var).pack(anchor="w", padx=6)

        # 指标优化调节
        af = ttk.LabelFrame(frame, text="指标优化调节（字段名 -> 因子，>1 提升、<1 降低）")
        af.pack(fill="both", expand=True, pady=4)
        ar = ttk.Frame(af); ar.pack(fill="x", padx=6, pady=4)
        ttk.Label(ar, text="字段名:").pack(side="left")
        self.adj_field_var = tk.StringVar()
        ttk.Entry(ar, textvariable=self.adj_field_var, width=12).pack(side="left", padx=4)
        ttk.Label(ar, text="因子:").pack(side="left", padx=(8, 0))
        self.adj_factor_var = tk.DoubleVar(value=1.0)
        ttk.Spinbox(ar, from_=0.1, to=10, increment=0.1, width=6,
                    textvariable=self.adj_factor_var).pack(side="left", padx=4)
        ttk.Button(ar, text="添加/更新", command=self._add_adjust).pack(side="left", padx=4)
        ttk.Button(ar, text="删除", command=self._del_adjust).pack(side="left")
        self.adj_tree = ttk.Treeview(af, columns=("字段", "因子"), show="headings", height=4)
        self.adj_tree.heading("字段", text="字段")
        self.adj_tree.heading("因子", text="因子")
        self.adj_tree.column("字段", width=150, anchor="center")
        self.adj_tree.column("因子", width=80, anchor="center")
        self.adj_tree.pack(fill="both", expand=True, padx=4, pady=4)

        # 零分处理
        zf = ttk.LabelFrame(frame, text="零分处理")
        zf.pack(fill="x", pady=4)
        self.zero_var = tk.StringVar(value="keep")
        for v, t in [("keep", "保留"), ("exclude", "排除"), ("avg", "按均值填充")]:
            ttk.Radiobutton(zf, text=t, variable=self.zero_var, value=v).pack(side="left", padx=8)

    def _add_adjust(self):
        f = self.adj_field_var.get().strip()
        if not f:
            return
        factor = self.adj_factor_var.get()
        self.config.metric_adjust[f] = factor
        for i in self.adj_tree.get_children(""):
            if self.adj_tree.set(i, "字段") == f:
                self.adj_tree.item(i, values=(f, factor))
                return
        self.adj_tree.insert("", "end", values=(f, factor))

    def _del_adjust(self):
        sel = self.adj_tree.selection()
        if not sel:
            return
        f = self.adj_tree.set(sel[0], "字段")
        self.adj_tree.delete(sel)
        self.config.metric_adjust.pop(f, None)

    def _gather_config(self):
        c = self.config
        c.num_classes = self.num_classes_var.get()
        c.seed = self.seed_var.get()
        c.w_total = self.weight_vars["w_total"].get()
        c.w_subject = self.weight_vars["w_subject"].get()
        c.w_segment = self.weight_vars["w_segment"].get()
        c.w_category = self.weight_vars["w_category"].get()
        c.w_pointwise = self.weight_vars["w_pointwise"].get()
        c.w_size = self.weight_vars["w_size"].get()
        c.tol_mean = self.tol_mean_var.get()
        c.tol_count = self.tol_count_var.get()
        c.enable_pointwise = self.pointwise_var.get()
        c.enable_name_pinyin = self.pinyin_var.get()
        c.enable_segment_auto = self.seg_auto_var.get()
        c.zero_score_strategy = self.zero_var.get()
        # custom_sizes 在 _on_run 中收集

    # ---------- 4. 执行分班 ----------
    def _build_run_tab(self):
        f = self.tab_run
        top = ttk.Frame(f); top.pack(fill="x", padx=8, pady=8)
        self.run_btn = ttk.Button(top, text="▶ 执行分班", command=self._on_run)
        self.run_btn.pack(side="left")
        self.cancel_btn = ttk.Button(top, text="⏹ 中止", command=self._on_cancel, state="disabled")
        self.cancel_btn.pack(side="left", padx=4)
        # 迭代次数输入框
        ttk.Label(top, text="迭代次数:").pack(side="left", padx=(16, 0))
        self.max_iter_var = tk.IntVar(value=300)
        ttk.Spinbox(top, from_=50, to=2000, width=6,
                    textvariable=self.max_iter_var).pack(side="left", padx=4)
        ttk.Button(top, text="📝 导出完整日志", command=self._on_export_log).pack(side="right")

        self.progress = ttk.Progressbar(f, orient="horizontal",
                                         mode="determinate", length=600)
        self.progress.pack(fill="x", padx=8, pady=4)
        self.progress_var = tk.StringVar(value="就绪")
        ttk.Label(f, textvariable=self.progress_var).pack()
        # 迭代次数说明
        iter_hint = ("💡 迭代次数：控制优化算法的最大迭代轮数。每轮执行定向换班、贪心换班、"
                     "终局微调等操作。次数越多越接近最优，但超过300次后边际收益递减。"
                     "建议范围：100-500。设置为0表示使用默认值300。")
        ttk.Label(f, text=iter_hint, foreground="#666666",
                  wraplength=900, justify="left").pack(fill="x", padx=8)

        # 单页概览：使用 Text + Scrollbar（可靠的可滚动方案）
        sf = ttk.LabelFrame(f, text="分班结果统计（完整概览）")
        sf.pack(fill="both", expand=True, padx=8, pady=4)
        text_frame = ttk.Frame(sf)
        text_frame.pack(fill="both", expand=True, padx=4, pady=4)
        vsb = ttk.Scrollbar(text_frame, orient="vertical")
        vsb.pack(side="right", fill="y")
        self.run_text = tk.Text(text_frame, font=("Consolas", 10), wrap="word",
                                yscrollcommand=vsb.set, state="normal")
        self.run_text.pack(side="left", fill="both", expand=True)
        vsb.config(command=self.run_text.yview)
        # 配置文本标签样式
        self.run_text.tag_configure("header", font=("Consolas", 11, "bold"),
                                     foreground="#0066cc", spacing1=4, spacing3=2)
        self.run_text.tag_configure("subheader", font=("Consolas", 10, "bold"),
                                     foreground="#333333", spacing1=2)
        self.run_text.tag_configure("group_hdr", font=("Consolas", 10, "bold"),
                                     foreground="#0066cc", background="#e8f0fe",
                                     spacing1=2, spacing3=2)
        self.run_text.tag_configure("warn", foreground="#cc0000")
        self.run_text.tag_configure("ok", foreground="#008800")
        self.run_text.tag_configure("blue", foreground="#0066cc")
        self.run_text.tag_configure("indent", lmargin1=20, lmargin2=20)

    def _on_cancel(self):
        self._cancel_flag = True
        self.progress_var.set("正在中止…")

    def _build_full_log_lines(self) -> list[tuple[str, str]]:
        """生成完整日志内容（文本行 + 标签），供界面显示和文件导出共用"""
        lines: list[tuple[str, str]] = []
        H = "header"
        S = "subheader"
        G = "group_hdr"
        I = "indent"
        W = "warn"
        OK = "ok"
        B = "blue"
        N = ""

        lines.append(("=" * 70, H))
        lines.append(("教务分班软件 - 完整日志", H))
        lines.append(("=" * 70, H))
        lines.append((f"时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", N))
        lines.append((f"数据文件: {self.excel_path}", N))
        lines.append((f"班级数: {self.config.num_classes}", N))
        lines.append((f"学生数: {len(self.students)}", N))
        lines.append(("-" * 70, N))
        lines.append(("", N))

        # ── 一、约束条件配置 ──
        lines.append(("【一、约束条件配置】", H))
        lines.append(("-" * 50, N))

        if self.config.class_types:
            lines.append(("1. 班型配置", S))
            for spec in self.config.class_types:
                cls_str = ",".join(str(c + 1) + "班" for c in spec.class_indices)
                line = f"  {spec.name}: {cls_str}"
                if spec.rank_floor is not None or spec.rank_ceil is not None:
                    if spec.rank_floor is not None and spec.rank_ceil is not None:
                        line += f"，第{spec.rank_floor}~{spec.rank_ceil}名"
                    elif spec.rank_floor is not None:
                        line += f"，第{spec.rank_floor}名及以后"
                    elif spec.rank_ceil is not None:
                        line += f"，前{spec.rank_ceil}名"
                if spec.score_floor is not None or spec.score_ceil is not None:
                    if spec.score_floor is not None and spec.score_ceil is not None:
                        line += f"，{spec.score_floor}~{spec.score_ceil}分"
                    elif spec.score_floor is not None:
                        line += f"，{spec.score_floor}分以上"
                    elif spec.score_ceil is not None:
                        line += f"，{spec.score_ceil}分以下"
                lines.append((line, I))
            if self.engine and self.engine._class_type_of:
                type_cnt = {}
                for t in self.engine._class_type_of.values():
                    type_cnt[t] = type_cnt.get(t, 0) + 1
                lines.append((f"  实际分配: {type_cnt}", I))
            lines.append(("", N))

        if self.constraints.bindings:
            lines.append(("2. 同班捆绑组", S))
            for i, g in enumerate(self.constraints.bindings, 1):
                names = [self.students[s].name for s in g.sids]
                target = f"→ {g.target_class + 1}班" if g.target_class is not None else ""
                lines.append((f"  {i}. {'、'.join(names)} {target}", I))
            lines.append(("", N))

        if self.constraints.mutexes:
            lines.append(("3. 互斥分散组", S))
            for i, g in enumerate(self.constraints.mutexes, 1):
                names = [self.students[s].name for s in g.sids]
                lines.append((f"  {i}. {'、'.join(names)}", I))
            lines.append(("", N))

        if self.config.segment_rules:
            lines.append(("4. 自定义分数段均衡规则", S))
            for i, rule in enumerate(self.config.segment_rules, 1):
                mode = "按排名" if rule.by_rank else "按分数"
                lines.append((f"  {i}. {rule.field_name} {mode}分段，切点: {rule.cuts}", I))
            lines.append(("", N))

        if self.config.enable_name_pinyin:
            lines.append(("5. 姓名同音检测：已启用", S))
            if self.constraints.name_conflicts:
                for i, g in enumerate(self.constraints.name_conflicts, 1):
                    names = [self.students[s].name for s in g.sids]
                    lines.append((f"  {i}. {'、'.join(names)}（读音相同或相近）", I))
            else:
                lines.append(("  未检测到同音姓名", I))
            lines.append(("", N))

        lines.append(("6. 其他配置", S))
        lines.append((f"  逐点均衡: {'启用' if self.config.enable_pointwise else '关闭'}", I))
        lines.append((f"  自动分段均衡: {'启用' if self.config.enable_segment_auto else '关闭'}", I))
        lines.append((f"  容差(人数): {self.config.tol_count}, 容差(均分): {self.config.tol_mean}", I))
        if self.config.custom_sizes and any(v is not None for v in self.config.custom_sizes):
            custom_str = ", ".join(f"{i+1}班={v}" for i, v in enumerate(self.config.custom_sizes) if v is not None)
            lines.append((f"  自定义班级人数: {custom_str}", I))
        lines.append(("", N))

        # ── 二、自动分班结果 ──
        lines.append(("【二、自动分班结果】", H))
        lines.append(("-" * 50, N))
        if self.stats:
            lines.append((f"迭代次数: {self.stats.iterations}", N))
            lines.append((f"是否收敛达标: {'是 ✓' if self.stats.converged else '否（已用尽迭代）'}", N))
            lines.append((f"最终成本: {self.stats.final_cost:.4f}", N))
            lines.append(("", N))
            lines.append(("统计指标(全局，跨所有班，含不同班型):", S))
            lines.append((f"  总分均分最大偏差: {self.stats.total_dev:.4f}", I))
            lines.append((f"  单科均分最大偏差: {self.stats.subject_dev:.4f}", I))
            lines.append((f"  分段人数最大偏差: {self.stats.segment_dev}", I))
            lines.append((f"  类别人数最大偏差: {self.stats.category_dev}", I))
            lines.append((f"  班级人数最大偏差: {self.stats.size_dev}", I))
            lines.append((f"  逐点偏差: {self.stats.pointwise_dev}", I))
            lines.append((f"  硬约束违反度: {self.stats.hard_penalty:.1f}", I))
            lines.append(("", N))
            if self.stats.group_stats:
                lines.append(("统计指标(各班型组内偏差，重点班/普通班分开计算):", S))
                for gs in self.stats.group_stats:
                    cls_str = ",".join(f"{c}班" for c in gs["classes"])
                    lines.append((f"  [{gs['name']}] {cls_str}:", G))
                    lines.append((f"    总分均分偏差: {gs['total_dev']:.4f}", I))
                    lines.append((f"    单科均分偏差: {gs['subject_dev']:.4f}", I))
                    lines.append((f"    分段人数偏差: {gs['segment_dev']}", I))
                    lines.append((f"    类别人数偏差: {gs['category_dev']}", I))
                    lines.append((f"    各班人数: {gs['counts']}", I))
                    means_str = " / ".join(f"{c}班={m:.2f}" for c, m in zip(gs["classes"], gs["total_means"]))
                    lines.append((f"    各班总分均分: {means_str}", I))
                lines.append(("", N))

        # 各班概况
        lines.append(("各班概况:", S))
        score_fields = [fd.name for fd in self.fields if fd.is_score]
        header = f"{'班级':<8}{'人数':<8}" + "".join(f"{f+'均分':<14}" for f in score_fields)
        lines.append((header, N))
        for cls in self.classes:
            line = f"{cls.name:<8}{len(cls):<8}"
            for f in score_fields:
                if cls.students:
                    avg = sum(s.score_of(f) for s in cls.students) / len(cls.students)
                    line += f"{avg:<14.2f}"
                else:
                    line += f"{'0.00':<14}"
            lines.append((line, N))
        lines.append(("", N))

        # ── 三、自动分段均衡 ──
        if self.classes:
            lines.append(("【三、自动分段均衡（高/中/低三段）】", H))
            lines.append(("-" * 50, N))
            if self.engine and hasattr(self.engine, '_group_seg_bounds') and self.engine._group_seg_bounds:
                groups = self.engine._class_type_groups()
                group_names = self.engine._group_names()
                for gi, group in enumerate(groups):
                    if gi in self.engine._group_seg_bounds:
                        c1, c2, total, hi_score, mid_score = self.engine._group_seg_bounds[gi]
                        gname = group_names[gi] if gi < len(group_names) else f"组{gi+1}"
                        cls_str = ",".join(f"{c+1}班" for c in group)
                        lines.append((f"  {gname}({cls_str}) 共{total}人:", I))
                        lines.append((f"    高分段: 第1~{c1}名 (≥{hi_score:.1f}分)", I))
                        lines.append((f"    中分段: 第{c1+1}~{c2}名 ({mid_score:.1f}~{hi_score:.1f}分)", I))
                        lines.append((f"    低分段: 第{c2+1}~{total}名 (≤{mid_score:.1f}分)", I))
            else:
                ordered = sorted(self.students, key=lambda s: -s.total_score)
                total = len(ordered)
                c1, c2 = total // 3, total * 2 // 3
                hi_score = ordered[c1 - 1].total_score if c1 > 0 else 0.0
                mid_score = ordered[c2 - 1].total_score if c2 > 0 else 0.0
                lines.append((f"  全局分段 共{total}人:", I))
                lines.append((f"    高分段: 第1~{c1}名 (≥{hi_score:.1f}分)", I))
                lines.append((f"    中分段: 第{c1+1}~{c2}名 ({mid_score:.1f}~{hi_score:.1f}分)", I))
                lines.append((f"    低分段: 第{c2+1}~{total}名 (≤{mid_score:.1f}分)", I))
            lines.append(("", N))
            # 各班分段人数表
            seg_of = self.engine._group_seg_of if (
                self.engine and hasattr(self.engine, '_group_seg_of') and self.engine._group_seg_of
            ) else None
            if seg_of is None:
                ordered = sorted(self.students, key=lambda s: -s.total_score)
                total = len(ordered)
                c1, c2 = total // 3, total * 2 // 3
                seg_of = {}
                for rank, st in enumerate(ordered, start=1):
                    seg_of[st.sid] = 0 if rank <= c1 else (1 if rank <= c2 else 2)
            lines.append((f"  {'班级':<8}{'高分段':<10}{'中分段':<10}{'低分段':<10}", S))
            for cls in self.classes:
                high = mid = low = 0
                for s in cls.students:
                    seg = seg_of.get(s.sid, 2)
                    if seg == 0: high += 1
                    elif seg == 1: mid += 1
                    else: low += 1
                lines.append((f"  {cls.name:<8}{high:<10}{mid:<10}{low:<10}", I))
            lines.append(("", N))

        # ── 四、自定义分数段各班人数 ──
        if self.config.segment_rules and self.classes:
            lines.append(("【四、自定义分数段各班人数】", H))
            lines.append(("-" * 50, N))
            ordered = sorted(self.students, key=lambda s: -s.total_score)
            rank_map = {st.sid: i + 1 for i, st in enumerate(ordered)}
            for rule in self.config.segment_rules:
                mode = "按排名" if rule.by_rank else "按分数"
                lines.append((f"字段: {rule.field_name}（{mode}分段，切点: {rule.cuts}）", S))
                cuts = [0] + rule.cuts + [99999]
                seg_names = [f"段{i+1}({cuts[i]}-{cuts[i+1]})" for i in range(len(cuts) - 1)]
                hdr = f"  {'班级':<8}"
                for sn in seg_names:
                    hdr += f"{sn:<18}"
                lines.append((hdr, N))
                for cls in self.classes:
                    seg_counts = {}
                    for s in cls.students:
                        if rule.by_rank:
                            rank = rank_map.get(s.sid, 0)
                            seg = rule.segment_of(0.0, rank=rank)
                        else:
                            seg = rule.segment_of(s.score_of(rule.field_name))
                        seg_counts[seg] = seg_counts.get(seg, 0) + 1
                    line = f"  {cls.name:<8}"
                    for i in range(len(seg_names)):
                        line += f"{seg_counts.get(i, 0):<18}"
                    lines.append((line, I))
                lines.append(("", N))

        # ── 五、手动换班日志 ──
        lines.append(("【五、手动换班日志】", H))
        lines.append(("-" * 50, N))
        if self.swap_log:
            lines.append((f"换班次数: {len(self.swap_log)}", N))
            lines.append(("", N))
            for i, entry in enumerate(self.swap_log, 1):
                lines.append((f"【{i}】", S))
                lines.append((entry, I))
                lines.append(("", N))
        else:
            lines.append(("暂无手动换班记录", I))
            lines.append(("", N))

        # ── 六、当前统计（手动调班后） ──
        if self.swap_log and self.stats:
            cur = self._compute_current_stats()
            lines.append(("【六、当前统计（手动调班后）】", H))
            lines.append(("-" * 50, N))
            lines.append((f"总分均分最大偏差: {cur.total_dev:.4f}", N))
            lines.append((f"单科均分最大偏差: {cur.subject_dev:.4f}", N))
            lines.append((f"分段人数最大偏差: {cur.segment_dev}", N))
            lines.append((f"类别人数最大偏差: {cur.category_dev}", N))
            lines.append((f"班级人数最大偏差: {cur.size_dev}", N))
            lines.append((f"当前成本: {cur.final_cost:.4f}", N))
            lines.append(("", N))

        # ── 约束验证 ──
        rep = self.constraints.validate(self.students, self.config.num_classes)
        lines.append(("【约束验证】", H))
        lines.append(("-" * 50, N))
        if rep.satisfied:
            lines.append(("结果: 全部满足 ✓", OK))
        else:
            lines.append(("结果: 存在违反 ⚠", W))
            for v in rep.violations[:10]:
                lines.append((f"  • {v}", I))

        return lines

    def _on_export_log(self):
        """导出完整日志到文本文件"""
        if not self.stats and not self.swap_log:
            messagebox.showwarning("提示", "暂无分班结果或换班日志")
            return
        path = filedialog.asksaveasfilename(
            title="保存完整日志",
            defaultextension=".txt",
            filetypes=[("文本文件", "*.txt"), ("所有文件", "*.*")])
        if not path:
            return
        try:
            lines = self._build_full_log_lines()
            content = "\n".join(text for text, _ in lines)
            with open(path, "w", encoding="utf-8") as f:
                f.write(content)
            messagebox.showinfo("成功", f"日志已导出到:\n{path}")
        except Exception as e:
            messagebox.showerror("导出失败", str(e))

    def _on_run(self):
        if not self.students:
            messagebox.showwarning("提示", "请先导入数据")
            return
        # 重置手动调班日志（新一轮自动分班）
        self.auto_stats_text = None
        self.swap_log = []
        self._gather_config()
        # 姓名同音检测
        if self.config.enable_name_pinyin:
            self.constraints.detect_name_conflicts(self.students)
        # 人数自定义：收集到 config.custom_sizes
        self.config.custom_sizes = []
        for i, v in enumerate(self.size_vars[:self.config.num_classes]):
            val = v.get().strip()
            if val:
                try:
                    self.config.custom_sizes.append(int(val))
                except ValueError:
                    self.config.custom_sizes.append(None)
            else:
                self.config.custom_sizes.append(None)
        # 迭代次数
        self.config.max_iterations = self.max_iter_var.get()

        self.run_btn.config(state="disabled")
        self.cancel_btn.config(state="normal")
        self.progress["value"] = 0
        self.progress_var.set("初始化…")
        self._cancel_flag = False

        def progress_cb(it, total, msg):
            if self._cancel_flag:
                return
            self.progress["maximum"] = total
            self.progress["value"] = it
            self.progress_var.set(f"[{it}/{total}] {msg}")

        def worker():
            try:
                self.engine = BalanceEngine(
                    self.students, self.config, self.constraints,
                    progress_cb=progress_cb)
                classes, stats = self.engine.run()
                self.classes = classes
                self.stats = stats
                self.adjuster = ManualAdjuster(
                    self.students, self.config, self.constraints, self.engine)
                # 显示统计
                self.root.after(0, lambda: self._show_stats(stats))
            except Exception as e:
                self.root.after(0, lambda: messagebox.showerror("错误", str(e)))
            finally:
                self.root.after(0, lambda: self.run_btn.config(state="normal"))
                self.root.after(0, lambda: self.cancel_btn.config(state="disabled"))

        t = threading.Thread(target=worker, daemon=True)
        t.start()

    def _format_stats_text(self, stats: EngineStats, title: str = "分班结果统计") -> str:
        """格式化统计文本（不直接写入控件，便于快照保存）"""
        lines: list[str] = []
        lines.append(f"═══ {title} ═══\n\n")
        # 班型配置信息（便于核对实际使用的配置）
        if self.config.class_types:
            lines.append("── 班型配置 ──\n")
            for spec in self.config.class_types:
                cls_str = ",".join(str(c + 1) + "班" for c in spec.class_indices)
                rank_str = ""
                if spec.rank_floor is not None and spec.rank_ceil is not None:
                    rank_str = f"第{spec.rank_floor}~{spec.rank_ceil}名"
                elif spec.rank_floor is not None:
                    rank_str = f"第{spec.rank_floor}名及以后"
                elif spec.rank_ceil is not None:
                    rank_str = f"前{spec.rank_ceil}名"
                else:
                    rank_str = "不限名次"
                score_str = ""
                if spec.score_floor is not None and spec.score_ceil is not None:
                    score_str = f"{spec.score_floor}~{spec.score_ceil}分"
                elif spec.score_floor is not None:
                    score_str = f"{spec.score_floor}分以上"
                elif spec.score_ceil is not None:
                    score_str = f"{spec.score_ceil}分以下"
                else:
                    score_str = "不限分数"
                lines.append(f"  {spec.name} → {cls_str} | {rank_str} | {score_str}\n")
            # 实际班型分配人数
            if self.engine and self.engine._class_type_of:
                type_cnt = {}
                for t in self.engine._class_type_of.values():
                    type_cnt[t] = type_cnt.get(t, 0) + 1
                lines.append(f"  实际分配: {type_cnt}\n")
            lines.append("\n")
        lines.append(f"迭代次数:        {stats.iterations}\n")
        lines.append(f"是否收敛达标:    {'是 ✓' if stats.converged else '否（已用尽迭代）'}\n")
        lines.append(f"最终成本:        {stats.final_cost:.4f}\n\n")
        lines.append(f"总分均分最大偏差(全局):  {stats.total_dev:.4f}\n")
        lines.append(f"单科均分最大偏差(全局):  {stats.subject_dev:.4f}\n")
        lines.append(f"分段人数最大偏差:  {stats.segment_dev}\n")
        lines.append(f"类别人数最大偏差:  {stats.category_dev}\n")
        lines.append(f"班级人数最大偏差:  {stats.size_dev}\n")
        lines.append(f"逐点偏差:          {stats.pointwise_dev}\n")
        lines.append(f"硬约束违反度:      {stats.hard_penalty:.1f}\n\n")
        # 班型组分别统计
        if stats.group_stats:
            lines.append("── 各班型组内偏差（重点班/普通班分别计算） ──\n\n")
            for gs in stats.group_stats:
                cls_str = ",".join(f"{c}班" for c in gs["classes"])
                lines.append(f"  [{gs['name']}] 班级: {cls_str}\n")
                lines.append(f"    总分均分偏差: {gs['total_dev']:.4f}\n")
                lines.append(f"    单科均分偏差: {gs['subject_dev']:.4f}\n")
                lines.append(f"    分段人数偏差: {gs['segment_dev']}\n")
                lines.append(f"    类别人数偏差: {gs['category_dev']}\n")
                lines.append(f"    各班人数: {gs['counts']}\n")
                means_str = " / ".join(f"{c}班={m:.2f}" for c, m in zip(gs["classes"], gs["total_means"]))
                lines.append(f"    各班总分均分: {means_str}\n\n")
        # 各班详情
        lines.append("── 各班概况 ──\n")
        score_fields = [fd.name for fd in self.fields if fd.is_score]
        header = f"{'班级':<6}{'人数':<6}" + "".join(f"{f+'均分':<12}" for f in score_fields)
        lines.append(header + "\n")
        for cls in self.classes:
            line = f"{cls.name:<6}{len(cls):<6}"
            for f in score_fields:
                if cls.students:
                    avg = sum(s.score_of(f) for s in cls.students) / len(cls.students)
                    line += f"{avg:<12.2f}"
            lines.append(line + "\n")
        # 类别均衡详情（性别等）
        cat_fields = [fd.name for fd in self.fields if fd.is_category]
        if cat_fields:
            lines.append("\n── 类别均衡详情 ──\n")
            from collections import Counter
            for cf in cat_fields:
                lines.append(f"\n  [{cf}]\n")
                all_vals = set()
                for cls in self.classes:
                    for s in cls.students:
                        all_vals.add(s.category_of(cf))
                all_vals = sorted(all_vals)
                hdr = f"    {'班级':<6}" + "".join(f"{v+'人数':<10}" for v in all_vals)
                lines.append(hdr + "\n")
                col_counts = {v: [] for v in all_vals}
                for cls in self.classes:
                    cnt = Counter(s.category_of(cf) for s in cls.students)
                    line = f"    {cls.name:<6}"
                    for v in all_vals:
                        c = cnt.get(v, 0)
                        col_counts[v].append(c)
                        line += f"{c:<10}"
                    lines.append(line + "\n")
                dev_line = f"    {'偏差':<6}"
                for v in all_vals:
                    d = max(col_counts[v]) - min(col_counts[v]) if col_counts[v] else 0
                    dev_line += f"{d:<10}"
                lines.append(dev_line + "\n")
        # 约束验证
        rep = self.constraints.validate(self.students, self.config.num_classes)
        lines.append(f"\n── 约束验证: {'全部满足 ✓' if rep.satisfied else '存在违反 ⚠'} ──\n")
        for v in rep.violations[:10]:
            lines.append(f"  • {v}\n")
        return "".join(lines)

    def _show_stats(self, stats: EngineStats):
        """自动分班完成后调用：保存快照并刷新显示"""
        self.progress_var.set(f"完成 - 迭代 {stats.iterations} 次")
        # 保存自动分班统计快照（仅首次）
        if self.auto_stats_text is None:
            self.auto_stats_text = self._format_stats_text(stats, "自动分班结果统计")
        self._refresh_stats_display()
        # 刷新调班页和导出预览
        self._refresh_adjust_tab()
        self._refresh_export_preview()
        self.status_var.set("分班完成 - 可前往 [手动调班] 或 [结果导出]")

    def _compute_current_stats(self) -> EngineStats:
        """根据当前学生分配重新计算统计（手动调班后）"""
        assignment = {st.sid: st.assigned_class for st in self.students
                      if st.assigned_class is not None}
        # 同步引擎统计缓存
        self.engine._resync_stats(assignment)
        stats = EngineStats()
        self.engine._fill_stats(assignment, stats)
        stats.final_cost = self.engine._total_cost(assignment)
        stats.iterations = self.stats.iterations if self.stats else 0
        stats.converged = self.stats.converged if self.stats else False
        return stats

    def _refresh_stats_display(self):
        """刷新执行分班页面的完整日志（与导出日志格式完全一致）"""
        if not hasattr(self, "run_text"):
            return
        t = self.run_text
        t.delete("1.0", "end")
        if not self.classes and not self.stats:
            t.insert("end", "（暂无分班结果，请先执行分班）\n", "warn")
            return
        for text, tag in self._build_full_log_lines():
            if tag:
                t.insert("end", text + "\n", tag)
            else:
                t.insert("end", text + "\n")

    # ---------- 5. 手动调班 ----------
    def _build_adjust_tab(self):
        f = self.tab_adjust
        top = ttk.Frame(f); top.pack(fill="x", padx=8, pady=6)
        self.adjust_status = tk.StringVar(value="")
        ttk.Label(top, textvariable=self.adjust_status).pack(side="left", padx=8)

        # 双栏：左 A 班 / 右 B 班，可选中换班
        body = ttk.Frame(f); body.pack(fill="both", expand=True, padx=8, pady=4)
        left = ttk.LabelFrame(body, text="班级 A")
        left.pack(side="left", fill="both", expand=True, padx=4)
        right = ttk.LabelFrame(body, text="班级 B")
        right.pack(side="left", fill="both", expand=True, padx=4)

        # 班级选择
        la = ttk.Frame(left); la.pack(fill="x", padx=4, pady=4)
        ttk.Label(la, text="选择:").pack(side="left")
        self.class_a_var = tk.StringVar()
        self.class_a_combo = ttk.Combobox(la, textvariable=self.class_a_var,
                                          state="readonly", width=10)
        self.class_a_combo.pack(side="left", padx=4)
        self.class_a_combo.bind("<<ComboboxSelected>>",
                                lambda e: self._refresh_adjust_lists())

        ra = ttk.Frame(right); ra.pack(fill="x", padx=4, pady=4)
        ttk.Label(ra, text="选择:").pack(side="left")
        self.class_b_var = tk.StringVar()
        self.class_b_combo = ttk.Combobox(ra, textvariable=self.class_b_var,
                                          state="readonly", width=10)
        self.class_b_combo.pack(side="left", padx=4)
        self.class_b_combo.bind("<<ComboboxSelected>>",
                                lambda e: self._refresh_adjust_lists())

        # 学生列表（含性别列）
        cols = ("sid", "name", "gender", "total", "preset")
        self.list_a = ttk.Treeview(left, columns=cols, show="headings", height=18)
        for c, t, w in [("sid", "序号", 50), ("name", "姓名", 90),
                        ("gender", "性别", 50), ("total", "总分", 70), ("preset", "预设", 50)]:
            self.list_a.heading(c, text=t); self.list_a.column(c, width=w, anchor="center")
        self.list_a.pack(fill="both", expand=True, padx=4, pady=4)

        self.list_b = ttk.Treeview(right, columns=cols, show="headings", height=18)
        for c, t, w in [("sid", "序号", 50), ("name", "姓名", 90),
                        ("gender", "性别", 50), ("total", "总分", 70), ("preset", "预设", 50)]:
            self.list_b.heading(c, text=t); self.list_b.column(c, width=w, anchor="center")
        self.list_b.pack(fill="both", expand=True, padx=4, pady=4)

        # 操作按钮
        op = ttk.Frame(f); op.pack(fill="x", padx=8, pady=4)
        ttk.Button(op, text="←→ 一对一换班",
                   command=self._on_swap_pair).pack(side="left")
        ttk.Button(op, text="→ 移到 B 班",
                   command=lambda: self._on_move("a2b")).pack(side="left", padx=4)
        ttk.Button(op, text="← 移到 A 班",
                   command=lambda: self._on_move("b2a")).pack(side="left")

    def _refresh_adjust_tab(self):
        if not self.classes:
            return
        names = [c.name for c in self.classes]
        self.class_a_combo["values"] = names
        self.class_b_combo["values"] = names
        if names:
            if not self.class_a_var.get():
                self.class_a_var.set(names[0])
            if not self.class_b_var.get():
                self.class_b_var.set(names[-1] if len(names) > 1 else names[0])
        self._refresh_adjust_lists()

    def _refresh_adjust_lists(self):
        for tree in (self.list_a, self.list_b):
            tree.delete(*tree.get_children())
        a_name = self.class_a_var.get()
        b_name = self.class_b_var.get()
        # 查找性别字段名（自动识别）
        gender_field = None
        for fd in self.fields:
            if fd.is_category and ("性" in fd.name or "别" in fd.name):
                gender_field = fd.name
                break
        for cls, tree in [(self._find_class(a_name), self.list_a),
                          (self._find_class(b_name), self.list_b)]:
            if not cls:
                continue
            # 按总分从高到低排序
            sorted_students = sorted(cls.students, key=lambda s: -s.total_score)
            for st in sorted_students:
                gender = st.category_of(gender_field) if gender_field else ""
                tree.insert("", "end", iid=str(st.sid), values=(
                    st.sid + 1, st.name, gender, round(st.total_score, 1),
                    st.preset_class or ""))

    def _find_class(self, name: str):
        for c in self.classes:
            if c.name == name:
                return c
        return None

    def _class_name_to_index(self, name: str) -> int:
        for i, c in enumerate(self.classes):
            if c.name == name:
                return i
        return -1

    def _get_class_type(self, sid: int) -> str:
        """获取学生的班型名（重点班/卓越班/普通班等）"""
        if self.engine and sid in self.engine._class_type_of:
            return self.engine._class_type_of[sid]
        return "普通班"

    def _constraint_desc(self, sid: int) -> str:
        """获取学生的约束类别描述（用于换班提示）"""
        parts = []
        st = self.students[sid]
        if st.preset_class is not None:
            parts.append(f"预设{st.preset_class}班")
        if self.constraints.binding_of(sid) is not None:
            parts.append("捆绑组")
        if self.constraints.mutex_groups_of(sid):
            parts.append("互斥组")
        if self.constraints.name_groups_of(sid):
            parts.append("姓名冲突")
        return "、".join(parts) if parts else "无"

    def _on_swap_pair(self):
        if not self.adjuster:
            return
        sa = self.list_a.selection()
        sb = self.list_b.selection()
        if not sa or not sb:
            messagebox.showwarning("提示", "请在 A/B 班级各选一名学生")
            return
        sid_a = int(sa[0]); sid_b = int(sb[0])
        # 检查班型差异 / 约束类别差异
        type_a = self._get_class_type(sid_a)
        type_b = self._get_class_type(sid_b)
        cons_a = self._constraint_desc(sid_a)
        cons_b = self._constraint_desc(sid_b)
        force = False
        warns = []
        if type_a != type_b:
            warns.append(f"班型不同: {self.students[sid_a].name}={type_a}, "
                         f"{self.students[sid_b].name}={type_b}")
        if cons_a != cons_b:
            warns.append(f"约束类别不同: {self.students[sid_a].name}={cons_a}, "
                         f"{self.students[sid_b].name}={cons_b}")
        # 始终弹出确认对话框
        info_lines = [
            f"换班操作确认",
            f"",
            f"  {self.students[sid_a].name} ({self.class_a_var.get()} → {self.class_b_var.get()})",
            f"    班型: {type_a}  约束: {cons_a}  总分: {self.students[sid_a].total_score:.1f}",
            f"",
            f"  {self.students[sid_b].name} ({self.class_b_var.get()} → {self.class_a_var.get()})",
            f"    班型: {type_b}  约束: {cons_b}  总分: {self.students[sid_b].total_score:.1f}",
        ]
        if warns:
            info_lines.insert(0, "⚠ 跨班型/约束差异提示")
            info_lines.append("")
            info_lines.extend(warns)
        info_lines.append("")
        info_lines.append("确认执行换班？")
        msg = "\n".join(info_lines)
        if not messagebox.askyesno("换班确认", msg, icon="warning" if warns else "question"):
            return
        if type_a != type_b or cons_a != cons_b:
            force = True
        r = self.adjuster.swap_pair(sid_a, sid_b, force_class_type=force)
        if r.success:
            self.adjust_status.set(
                f"换班成功 - 成本: {r.old_cost:.4f} → {r.new_cost:.4f}")
            # 重建班级、追加日志、刷新统计
            self._post_manual_adjust(
                f"换班: {self.students[sid_a].name}({type_a}) ↔ "
                f"{self.students[sid_b].name}({type_b})",
                r.old_cost, r.new_cost)
        else:
            messagebox.showwarning("换班失败", r.message)

    def _on_move(self, direction: str):
        if not self.adjuster:
            return
        if direction == "a2b":
            sel = self.list_a.selection()
            target = self._class_name_to_index(self.class_b_var.get())
            dst_name = self.class_b_var.get()
            src_name = self.class_a_var.get()
        else:
            sel = self.list_b.selection()
            target = self._class_name_to_index(self.class_a_var.get())
            dst_name = self.class_a_var.get()
            src_name = self.class_b_var.get()
        if not sel or target < 0:
            return
        sid = int(sel[0])
        # 检查班型差异 / 约束类别差异
        src_class = self.students[sid].assigned_class or 0
        type_src = self._get_class_type(sid)
        # 目标班的班型：查找该班是否有班型定义
        type_dst = "普通班"
        if self.engine:
            for ct_name, allowed in (
                (v, self.engine._class_type_filter[k]) for k, v in self.engine._class_type_of.items()
            ):
                if target in allowed:
                    type_dst = ct_name
                    break
        cons_src = self._constraint_desc(sid)
        force = False
        # 始终弹出确认对话框
        info_lines = [
            f"移班操作确认",
            f"",
            f"  学生: {self.students[sid].name}",
            f"  来源: {src_name}  →  目标: {dst_name}",
            f"  班型: {type_src} → {type_dst}",
            f"  约束: {cons_src}",
            f"  总分: {self.students[sid].total_score:.1f}",
        ]
        if type_src != type_dst:
            info_lines.insert(0, "⚠ 跨班型移班提示")
        info_lines.append("")
        info_lines.append("确认执行移班？")
        msg = "\n".join(info_lines)
        if not messagebox.askyesno("移班确认", msg, icon="warning" if type_src != type_dst else "question"):
            return
        if type_src != type_dst:
            force = True
        r = self.adjuster.move_one(sid, target, force_class_type=force)
        if r.success:
            self.adjust_status.set(
                f"移班成功 - 成本: {r.old_cost:.4f} → {r.new_cost:.4f}")
            self._post_manual_adjust(
                f"移班: {self.students[sid].name} {src_name}→{dst_name}",
                r.old_cost, r.new_cost)
        else:
            messagebox.showwarning("移班失败", r.message)

    def _post_manual_adjust(self, action_desc: str, old_cost: float, new_cost: float):
        """手动调班后：重建班级、追加日志（含完整统计）、刷新显示"""
        # 重建班级列表（学生 assigned_class 已更新）
        assignment = {st.sid: st.assigned_class for st in self.students
                      if st.assigned_class is not None}
        self.classes = self.engine._build_classes(assignment)
        # 追加换班日志（包含详细统计：各班人数、总分均分、各科均分）
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        delta = new_cost - old_cost
        arrow = "↑" if delta > 0 else ("↓" if delta < 0 else "→")
        cost_str = f"{old_cost:.4f}→{new_cost:.4f} ({arrow}{abs(delta):.4f})"
        # 生成各班完整均分摘要（总分 + 各单科）
        score_fields = [fd.name for fd in self.fields if fd.is_score]
        avg_items = []
        for cls in self.classes:
            if cls.students:
                parts = [f"{cls.name}({len(cls)}人)"]
                total_avg = sum(s.score_of("总分") for s in cls.students) / len(cls.students)
                parts.append(f"总分={total_avg:.1f}")
                for f in score_fields:
                    if f == "总分":
                        continue
                    f_avg = sum(s.score_of(f) for s in cls.students) / len(cls.students)
                    parts.append(f"{f}={f_avg:.1f}")
                avg_items.append(" ".join(parts))
        avg_summary = " | ".join(avg_items)
        # 格式：时间|操作|成本变化|均分摘要（用于表格显示）
        table_entry = f"{ts}|{action_desc}|{cost_str}|{avg_summary}"
        self.swap_log.append(table_entry)
        # 刷新统计区（保留自动分班快照 + 追加日志）
        self._refresh_stats_display()
        # 刷新调班列表、班级统计、导出预览
        self._refresh_adjust_lists()
        self._refresh_stats_tab()
        self._refresh_export_preview()

    def _on_reoptimize(self):
        if not self.adjuster:
            return
        stats = self.adjuster.reoptimize()
        # 重建班级（reoptimize 修改了 assignment 但未重建 classes）
        assignment = {st.sid: st.assigned_class for st in self.students
                      if st.assigned_class is not None}
        self.classes = self.engine._build_classes(assignment)
        # 记录到日志
        ts = datetime.now().strftime("%H:%M:%S")
        self.swap_log.append(f"[{ts}] 重新优化（保持手动调整）  成本: {stats.final_cost:.4f}")
        self._refresh_stats_display()
        self._refresh_adjust_lists()
        self._refresh_stats_tab()
        self._refresh_export_preview()

    # ---------- 6. 班级统计 ----------
    def _build_stats_tab(self):
        """班级统计页面：使用多个 Treeview 表格展示分班结果"""
        f = self.tab_stats
        top = ttk.Frame(f); top.pack(fill="x", padx=8, pady=6)
        ttk.Label(top, text="班级统计（每次调班后自动刷新最新结果）",
                  font=("Microsoft YaHei UI", 10, "bold")).pack(side="left")
        self.stats_status = tk.StringVar(value="")
        ttk.Label(top, textvariable=self.stats_status, foreground="blue").pack(side="left", padx=10)
        ttk.Button(top, text="💾 导出统计表为Excel",
                   command=self._on_export_stats_excel).pack(side="right")

        # 使用 Notebook 组织多个表格
        nb = ttk.Notebook(f)
        nb.pack(fill="both", expand=True, padx=8, pady=4)

        # 表格1：各科平均分
        f1 = ttk.Frame(nb); nb.add(f1, text="各科平均分")
        sf1 = ttk.LabelFrame(f1, text="各班各科平均分")
        sf1.pack(fill="both", expand=True, padx=4, pady=4)
        self.stats_scores_tree = ttk.Treeview(sf1, show="headings", height=15)
        vsb1 = ttk.Scrollbar(sf1, orient="vertical", command=self.stats_scores_tree.yview)
        self.stats_scores_tree.configure(yscrollcommand=vsb1.set)
        self.stats_scores_tree.pack(side="left", fill="both", expand=True, padx=4, pady=4)
        vsb1.pack(side="right", fill="y")

        # 表格2：自动分段均衡
        f2 = ttk.Frame(nb); nb.add(f2, text="自动分段均衡")
        sf2 = ttk.LabelFrame(f2, text="自动分段均衡（高/中/低三段，前1/3、中1/3、后1/3）")
        sf2.pack(fill="both", expand=True, padx=4, pady=4)
        self.stats_seg_auto_text = tk.Text(sf2, height=4, font=("Consolas", 10), wrap="word")
        self.stats_seg_auto_text.pack(fill="x", padx=4, pady=2)
        self.stats_seg_auto_tree = ttk.Treeview(sf2, show="headings", height=12)
        vsb2 = ttk.Scrollbar(sf2, orient="vertical", command=self.stats_seg_auto_tree.yview)
        self.stats_seg_auto_tree.configure(yscrollcommand=vsb2.set)
        self.stats_seg_auto_tree.pack(side="left", fill="both", expand=True, padx=4, pady=4)
        vsb2.pack(side="right", fill="y")

        # 表格3：自定义分数段
        f3 = ttk.Frame(nb); nb.add(f3, text="自定义分数段")
        sf3 = ttk.LabelFrame(f3, text="自定义分数段各班人数")
        sf3.pack(fill="both", expand=True, padx=4, pady=4)
        self.stats_seg_custom_tree = ttk.Treeview(sf3, show="headings", height=15)
        vsb3 = ttk.Scrollbar(sf3, orient="vertical", command=self.stats_seg_custom_tree.yview)
        self.stats_seg_custom_tree.configure(yscrollcommand=vsb3.set)
        self.stats_seg_custom_tree.pack(side="left", fill="both", expand=True, padx=4, pady=4)
        vsb3.pack(side="right", fill="y")

        # 表格4：类别均衡
        f4 = ttk.Frame(nb); nb.add(f4, text="类别均衡")
        sf4 = ttk.LabelFrame(f4, text="类别均衡（性别/民族等）")
        sf4.pack(fill="both", expand=True, padx=4, pady=4)
        self.stats_category_tree = ttk.Treeview(sf4, show="headings", height=15)
        vsb4 = ttk.Scrollbar(sf4, orient="vertical", command=self.stats_category_tree.yview)
        self.stats_category_tree.configure(yscrollcommand=vsb4.set)
        self.stats_category_tree.pack(side="left", fill="both", expand=True, padx=4, pady=4)
        vsb4.pack(side="right", fill="y")

    def _refresh_stats_tab(self):
        """刷新班级统计页面：动态计算最新结果并填充 Treeview"""
        if not hasattr(self, "stats_scores_tree"):
            return
        if not self.classes:
            self.stats_status.set("（暂无分班结果）")
            return

        cur = self._compute_current_stats()
        self.stats_status.set(
            f"当前成本: {cur.final_cost:.4f} | 总分偏差: {cur.total_dev:.4f} | "
            f"班级人数偏差: {cur.size_dev}")

        score_fields = [fd.name for fd in self.fields if fd.is_score]

        # ════════════ 表格1：各科平均分 ════════════
        tree = self.stats_scores_tree
        for item in tree.get_children():
            tree.delete(item)
        cols = ["班级", "人数"] + [f + "均分" for f in score_fields]
        tree["columns"] = cols
        for c in cols:
            tree.heading(c, text=c)
            tree.column(c, width=90, anchor="center")
        for cls in self.classes:
            vals = [cls.name, len(cls)]
            for f in score_fields:
                if cls.students:
                    avg = sum(s.score_of(f) for s in cls.students) / len(cls.students)
                    vals.append(f"{avg:.2f}")
                else:
                    vals.append("0.00")
            tree.insert("", "end", values=vals)
        # 合计行
        total_vals = ["合计", sum(len(c) for c in self.classes)]
        for f in score_fields:
            total_sum = sum(s.score_of(f) for s in self.students if s.assigned_class is not None)
            total_cnt = sum(1 for s in self.students if s.assigned_class is not None)
            total_vals.append(f"{total_sum/total_cnt:.2f}" if total_cnt else "0.00")
        tree.insert("", "end", values=total_vals, tags=("total",))
        tree.tag_configure("total", background="#e8f0fe", font=("Microsoft YaHei UI", 9, "bold"))

        # ════════════ 表格2：自动分段均衡 ════════════
        self.stats_seg_auto_text.delete("1.0", "end")
        range_lines = []
        if self.engine and hasattr(self.engine, '_group_seg_bounds') and self.engine._group_seg_bounds:
            groups = self.engine._class_type_groups()
            group_names = self.engine._group_names()
            for gi, group in enumerate(groups):
                if gi in self.engine._group_seg_bounds:
                    c1, c2, total, hi_score, mid_score = self.engine._group_seg_bounds[gi]
                    gname = group_names[gi] if gi < len(group_names) else f"组{gi+1}"
                    cls_str = ",".join(f"{c+1}班" for c in group)
                    range_lines.append(
                        f"{gname}({cls_str}) 共{total}人: "
                        f"高分段(第1~{c1}名, ≥{hi_score:.1f}分) | "
                        f"中分段(第{c1+1}~{c2}名, {mid_score:.1f}~{hi_score:.1f}分) | "
                        f"低分段(第{c2+1}~{total}名, ≤{mid_score:.1f}分)")
        else:
            ordered = sorted(self.students, key=lambda s: -s.total_score)
            total = len(ordered)
            c1, c2 = total // 3, total * 2 // 3
            hi_score = ordered[c1 - 1].total_score if c1 > 0 else 0.0
            mid_score = ordered[c2 - 1].total_score if c2 > 0 else 0.0
            range_lines.append(
                f"全局分段 共{total}人: "
                f"高分段(第1~{c1}名, ≥{hi_score:.1f}分) | "
                f"中分段(第{c1+1}~{c2}名, {mid_score:.1f}~{hi_score:.1f}分) | "
                f"低分段(第{c2+1}~{total}名, ≤{mid_score:.1f}分)")
        self.stats_seg_auto_text.insert("end", "\n".join(range_lines))

        tree2 = self.stats_seg_auto_tree
        for item in tree2.get_children():
            tree2.delete(item)
        cols2 = ["班级", "高分段", "中分段", "低分段"]
        tree2["columns"] = cols2
        for c in cols2:
            tree2.heading(c, text=c)
            tree2.column(c, width=100, anchor="center")
        seg_of = self.engine._group_seg_of if (
            self.engine and hasattr(self.engine, '_group_seg_of') and self.engine._group_seg_of
        ) else None
        if seg_of is None:
            ordered = sorted(self.students, key=lambda s: -s.total_score)
            total = len(ordered)
            c1, c2 = total // 3, total * 2 // 3
            seg_of = {}
            for rank, st in enumerate(ordered, start=1):
                seg_of[st.sid] = 0 if rank <= c1 else (1 if rank <= c2 else 2)
        for cls in self.classes:
            high = mid = low = 0
            for s in cls.students:
                seg = seg_of.get(s.sid, 2)
                if seg == 0: high += 1
                elif seg == 1: mid += 1
                else: low += 1
            tree2.insert("", "end", values=(cls.name, high, mid, low))

        # ════════════ 表格3：自定义分数段 ════════════
        tree3 = self.stats_seg_custom_tree
        for item in tree3.get_children():
            tree3.delete(item)
        if self.config.segment_rules and self.classes:
            ordered = sorted(self.students, key=lambda s: -s.total_score)
            rank_map = {st.sid: i + 1 for i, st in enumerate(ordered)}
            for rule in self.config.segment_rules:
                mode = "按排名" if rule.by_rank else "按分数"
                cuts = [0] + rule.cuts + [99999]
                seg_names = [f"段{i+1}({cuts[i]}-{cuts[i+1]})" for i in range(len(cuts) - 1)]
                cols3 = ["班级"] + seg_names
                tree3["columns"] = cols3
                for c in cols3:
                    tree3.heading(c, text=c)
                    tree3.column(c, width=120, anchor="center")
                for cls in self.classes:
                    seg_counts = {}
                    for s in cls.students:
                        if rule.by_rank:
                            rank = rank_map.get(s.sid, 0)
                            seg = rule.segment_of(0.0, rank=rank)
                        else:
                            seg = rule.segment_of(s.score_of(rule.field_name))
                        seg_counts[seg] = seg_counts.get(seg, 0) + 1
                    vals = [cls.name] + [seg_counts.get(i, 0) for i in range(len(seg_names))]
                    tree3.insert("", "end", values=vals)
        else:
            tree3["columns"] = ["提示"]
            tree3.heading("提示", text="提示")
            tree3.column("提示", width=300, anchor="center")
            tree3.insert("", "end", values=("未配置自定义分数段规则",))

        # ════════════ 表格4：类别均衡 ════════════
        tree4 = self.stats_category_tree
        for item in tree4.get_children():
            tree4.delete(item)
        cat_fields = [fd.name for fd in self.fields if fd.is_category]
        if cat_fields and self.classes:
            from collections import Counter
            all_vals = []
            for cf in cat_fields:
                vals = set()
                for cls in self.classes:
                    for s in cls.students:
                        vals.add(s.category_of(cf))
                all_vals.extend([(cf, v) for v in sorted(vals)])
            col_labels = [f"{cf}-{v}" for cf, v in all_vals]
            cols4 = ["班级"] + col_labels + ["偏差"]
            tree4["columns"] = cols4
            for c in cols4:
                tree4.heading(c, text=c)
                tree4.column(c, width=100, anchor="center")
            row_counts = []
            for cls in self.classes:
                vals = [cls.name]
                row_cnt = []
                for cf, v in all_vals:
                    cnt = Counter(s.category_of(cf) for s in cls.students)
                    c = cnt.get(v, 0)
                    vals.append(str(c))
                    row_cnt.append((cf, v, c))
                row_counts.append(row_cnt)
                if row_cnt:
                    max_c = max(c for _, _, c in row_cnt)
                    min_c = min(c for _, _, c in row_cnt)
                    vals.append(str(max_c - min_c))
                else:
                    vals.append("0")
                tree4.insert("", "end", values=vals)
            # 合计行
            total_vals = ["合计"]
            for cf, v in all_vals:
                total = sum(rc[i][2] for i, rc in enumerate(row_counts))
                total_vals.append(str(total))
            total_vals.append("")
            tree4.insert("", "end", values=total_vals, tags=("total",))
            tree4.tag_configure("total", background="#e8f0fe", font=("Microsoft YaHei UI", 9, "bold"))
        else:
            tree4["columns"] = ["提示"]
            tree4.heading("提示", text="提示")
            tree4.column("提示", width=300, anchor="center")
            tree4.insert("", "end", values=("无类别字段数据",))

    def _on_export_stats_excel(self):
        """导出班级统计为 Excel"""
        if not self.classes:
            messagebox.showwarning("提示", "暂无分班结果")
            return
        path = filedialog.asksaveasfilename(
            title="导出班级统计",
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx")],
            initialfile="班级统计.xlsx")
        if not path:
            return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            wb = Workbook()
            wb.remove(wb.active)
            score_fields = [fd.name for fd in self.fields if fd.is_score]
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
            center = Alignment(horizontal="center", vertical="center")

            # Sheet1: 各科平均分
            ws1 = wb.create_sheet("各科平均分")
            cols = ["班级", "人数"] + [f + "均分" for f in score_fields]
            for ci, c in enumerate(cols, 1):
                cell = ws1.cell(row=1, column=ci, value=c)
                cell.font = header_font; cell.fill = header_fill; cell.alignment = center
            for ri, cls in enumerate(self.classes, 2):
                ws1.cell(row=ri, column=1, value=cls.name).alignment = center
                ws1.cell(row=ri, column=2, value=len(cls)).alignment = center
                for fi, f in enumerate(score_fields, 3):
                    if cls.students:
                        avg = sum(s.score_of(f) for s in cls.students) / len(cls.students)
                        ws1.cell(row=ri, column=fi, value=round(avg, 2)).alignment = center
                    else:
                        ws1.cell(row=ri, column=fi, value=0.0).alignment = center
            for ci in range(1, len(cols) + 1):
                ws1.column_dimensions[chr(64 + ci) if ci <= 26 else "A" + chr(64 + ci - 26)].width = 12

            # Sheet2: 自动分段均衡
            ws2 = wb.create_sheet("自动分段均衡")
            cols2 = ["班级", "高分段", "中分段", "低分段"]
            for ci, c in enumerate(cols2, 1):
                cell = ws2.cell(row=1, column=ci, value=c)
                cell.font = header_font; cell.fill = header_fill; cell.alignment = center
            seg_of = self.engine._group_seg_of if (
                self.engine and hasattr(self.engine, '_group_seg_of') and self.engine._group_seg_of
            ) else None
            if seg_of is None:
                ordered = sorted(self.students, key=lambda s: -s.total_score)
                total = len(ordered)
                c1, c2 = total // 3, total * 2 // 3
                seg_of = {}
                for rank, st in enumerate(ordered, start=1):
                    seg_of[st.sid] = 0 if rank <= c1 else (1 if rank <= c2 else 2)
            for ri, cls in enumerate(self.classes, 2):
                high = mid = low = 0
                for s in cls.students:
                    seg = seg_of.get(s.sid, 2)
                    if seg == 0: high += 1
                    elif seg == 1: mid += 1
                    else: low += 1
                ws2.cell(row=ri, column=1, value=cls.name).alignment = center
                ws2.cell(row=ri, column=2, value=high).alignment = center
                ws2.cell(row=ri, column=3, value=mid).alignment = center
                ws2.cell(row=ri, column=4, value=low).alignment = center
            for ci in range(1, 5):
                ws2.column_dimensions[chr(64 + ci)].width = 12

            # Sheet3: 自定义分数段
            if self.config.segment_rules:
                ws3 = wb.create_sheet("自定义分数段")
                ordered = sorted(self.students, key=lambda s: -s.total_score)
                rank_map = {st.sid: i + 1 for i, st in enumerate(ordered)}
                for rule_idx, rule in enumerate(self.config.segment_rules):
                    mode = "按排名" if rule.by_rank else "按分数"
                    ws3.cell(row=1, column=1,
                             value=f"{rule.field_name}（{mode}分段，切点: {rule.cuts}）")
                    ws3.cell(row=1, column=1).font = header_font
                    cuts = [0] + rule.cuts + [99999]
                    seg_names = [f"段{i+1}({cuts[i]}-{cuts[i+1]})" for i in range(len(cuts) - 1)]
                    cols3 = ["班级"] + seg_names
                    for ci, c in enumerate(cols3, 1):
                        cell = ws3.cell(row=2, column=ci, value=c)
                        cell.font = header_font; cell.fill = header_fill; cell.alignment = center
                    for ri, cls in enumerate(self.classes, 3):
                        ws3.cell(row=ri, column=1, value=cls.name).alignment = center
                        for s in cls.students:
                            if rule.by_rank:
                                rank = rank_map.get(s.sid, 0)
                                seg = rule.segment_of(0.0, rank=rank)
                            else:
                                seg = rule.segment_of(s.score_of(rule.field_name))
                            seg_counts = {}
                            for s2 in cls.students:
                                if rule.by_rank:
                                    rank2 = rank_map.get(s2.sid, 0)
                                    seg2 = rule.segment_of(0.0, rank=rank2)
                                else:
                                    seg2 = rule.segment_of(s2.score_of(rule.field_name))
                                seg_counts[seg2] = seg_counts.get(seg2, 0) + 1
                            for si in range(len(seg_names)):
                                ws3.cell(row=ri, column=si + 2,
                                         value=seg_counts.get(si, 0)).alignment = center
                    for ci in range(1, len(cols3) + 1):
                        col_letter = chr(64 + ci) if ci <= 26 else "A" + chr(64 + ci - 26)
                        ws3.column_dimensions[col_letter].width = 14

            # Sheet4: 类别均衡
            cat_fields = [fd.name for fd in self.fields if fd.is_category]
            if cat_fields:
                ws4 = wb.create_sheet("类别均衡")
                from collections import Counter
                all_vals = []
                for cf in cat_fields:
                    vals = set()
                    for cls in self.classes:
                        for s in cls.students:
                            vals.add(s.category_of(cf))
                    all_vals.extend([(cf, v) for v in sorted(vals)])
                col_labels = [f"{cf}-{v}" for cf, v in all_vals]
                cols4 = ["班级"] + col_labels + ["偏差"]
                for ci, c in enumerate(cols4, 1):
                    cell = ws4.cell(row=1, column=ci, value=c)
                    cell.font = header_font; cell.fill = header_fill; cell.alignment = center
                for ri, cls in enumerate(self.classes, 2):
                    ws4.cell(row=ri, column=1, value=cls.name).alignment = center
                    row_cnt = []
                    for ci, (cf, v) in enumerate(all_vals, 2):
                        cnt = Counter(s.category_of(cf) for s in cls.students)
                        c = cnt.get(v, 0)
                        ws4.cell(row=ri, column=ci, value=c).alignment = center
                        row_cnt.append((cf, v, c))
                    if row_cnt:
                        max_c = max(c for _, _, c in row_cnt)
                        min_c = min(c for _, _, c in row_cnt)
                        ws4.cell(row=ri, column=len(all_vals) + 2,
                                 value=max_c - min_c).alignment = center
                    else:
                        ws4.cell(row=ri, column=len(all_vals) + 2, value=0).alignment = center
                for ci in range(1, len(cols4) + 1):
                    col_letter = chr(64 + ci) if ci <= 26 else "A" + chr(64 + ci - 26)
                    ws4.column_dimensions[col_letter].width = 12

            wb.save(path)
            messagebox.showinfo("导出成功", f"统计表已导出到:\n{path}")
        except Exception as e:
            messagebox.showerror("导出失败", str(e))

    # ---------- 7. 结果导出 ----------
    def _build_export_tab(self):
        f = self.tab_export
        top = ttk.Frame(f); top.pack(fill="x", padx=8, pady=8)
        ttk.Button(top, text="💾 导出 Excel",
                   command=self._on_export).pack(side="left")
        self.export_path_var = tk.StringVar(value="未导出")
        ttk.Label(top, textvariable=self.export_path_var).pack(side="left", padx=10)
        ttk.Label(top, text="(进入此页自动显示分班结果，无需先导出)",
                  foreground="gray").pack(side="left", padx=10)

        # 预览：可滚动，显示完整分班名单
        pf = ttk.LabelFrame(f, text="分班结果预览（自动显示）")
        pf.pack(fill="both", expand=True, padx=8, pady=4)
        cols = ("sid", "name", "class", "total")
        self.export_tree = ttk.Treeview(pf, columns=cols, show="headings")
        for c, t, w in [("sid", "序号", 60), ("name", "姓名", 120),
                        ("class", "班级", 80), ("total", "总分", 80)]:
            self.export_tree.heading(c, text=t)
            self.export_tree.column(c, width=w, anchor="center")
        # 添加滚动条
        vsb = ttk.Scrollbar(pf, orient="vertical", command=self.export_tree.yview)
        self.export_tree.configure(yscrollcommand=vsb.set)
        self.export_tree.pack(side="left", fill="both", expand=True, padx=4, pady=4)
        vsb.pack(side="right", fill="y")
        # 统计摘要
        self.export_summary_var = tk.StringVar(value="")
        ttk.Label(f, textvariable=self.export_summary_var,
                  foreground="blue").pack(fill="x", padx=8, pady=4)

    def _on_export(self):
        if not self.classes:
            messagebox.showwarning("提示", "请先执行分班")
            return
        path = filedialog.asksaveasfilename(
            title="保存分班结果",
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx"), ("所有文件", "*.*")])
        if not path:
            return
        try:
            saved = export_classes(path, self.students, self.fields,
                                   self.classes, self.teachers)
            self.export_path_var.set(saved)
            messagebox.showinfo("成功", f"已导出到:\n{saved}")
            # 刷新预览
            self._refresh_export_preview()
        except Exception as e:
            messagebox.showerror("导出失败", str(e))

    def _refresh_export_preview(self):
        """刷新导出页的预览名单（分班完成后自动调用，无需手动导出）"""
        if not hasattr(self, "export_tree"):
            return
        self.export_tree.delete(*self.export_tree.get_children())
        if not self.students:
            return
        for st in sorted(self.students, key=lambda s: (s.assigned_class or 0, -s.total_score)):
            self.export_tree.insert("", "end", values=(
                st.sid + 1, st.name,
                f"{st.assigned_class+1}班" if st.assigned_class is not None else "",
                round(st.total_score, 1)))
        # 统计摘要
        if hasattr(self, "export_summary_var") and self.classes:
            summary_parts = [f"共 {len(self.students)} 名学生 / {len(self.classes)} 个班级"]
            for cls in self.classes:
                if cls.students:
                    avg = sum(s.total_score for s in cls.students) / len(cls.students)
                    summary_parts.append(f"{cls.name}({len(cls)}人,均分{avg:.1f})")
            self.export_summary_var.set(" | ".join(summary_parts))


def main():
    root = tk.Tk()
    App(root)
    # 窗口居中
    root.update_idletasks()
    w = root.winfo_width()
    h = root.winfo_height()
    sw = root.winfo_screenwidth()
    sh = root.winfo_screenheight()
    x = max(0, (sw - w) // 2)
    y = max(0, (sh - h) // 2 - 20)
    root.geometry(f"+{x}+{y}")
    root.mainloop()


if __name__ == "__main__":
    main()
